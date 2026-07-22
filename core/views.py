import json
import logging
import threading
from collections import defaultdict
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.views import LoginView, LogoutView
from django.core.exceptions import PermissionDenied
import datetime
from django.db import transaction
from django.db.models import (
    Case,
    Count,
    IntegerField,
    OuterRef,
    Prefetch,
    Q,
    Subquery,
    When,
)
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.generic import DetailView, FormView, ListView, TemplateView

from core.forms import (
    DESPACHO_VALUE,
    EncaminhamentoForm,
    ImovelForm,
    ISICForm,
    OSEncerramentoForm,
    OSForm,
    OSVincularProcessoForm,
    OSWizardPasso2Form,
    OSWizardPasso2RelacionadoForm,
    OSWizardPasso3Form,
    ProducaoForm,
    RelatorioProducaoForm,
    RelatorioTempoRegistroForm,
    SiatUploadForm,
)
from core.relatorios import (
    exportar_producao_excel,
    exportar_tempo_registro_excel,
    linhas_relatorio_producao,
    linhas_relatorio_tempo_registro,
    relatorio_producao_por_servidor,
    relatorio_tempo_registro_processo,
)
from core.middleware import obter_vinculo_unidade_ativo
from core.mixins import (
    RequerAdminMixin,
    RequerHomologarMixin,
    RequerLoginJSONMixin,
    RequerLoginMixin,
)
from core.os_service import (
    _atualizar_status_unidade_encaminhamento,
    ativar_atendimento_interno_se_necessario,
    data_entrada_unidade,
    macroetapa_atual_os,
    origem_encaminhamento,
    os_ativas_por_unidade,
    os_da_unidade_atual,
    is_primeiro_encaminhamento,
    os_editavel_para_usuario,
    queryset_os_com_macroetapa,
    registrar_em_atendimento_na_unidade,
    registrar_encaminhamento_automatico,
    timeline_os,
    unidade_atual_da_os,
    _queryset_os_nao_encerradas,
)
from core import siat_index
from core.siat_config import SIAT_ARQUIVO_PATH
from core.siat_service import (
    atualizar_inscricao_do_arquivo,
    buscar_bloco_no_arquivo,
    buscar_inscricao_no_arquivo,
    buscar_por_logradouro_no_arquivo,
    carregar_arquivo_siat,
    contar_imoveis_siat_orfaos,
    limpar_imoveis_siat_orfaos,
    obter_coordenadas_bloco,
    vincular_imovel_a_os,
    vincular_isic_a_os,
    obter_status_arquivo_siat,
)
from core.models import (
    Comentario,
    Encaminhamento,
    Finalidade,
    Imovel,
    Natureza,
    OS,
    OsImovel,
    OsProcesso,
    OsUnidadeStatus,
    PreferenciaGerencial,
    ProcessoSei,
    LogAuditoria,
    Producao,
    ProducaoImovel,
    ProducaoStatusLog,
    Servidor,
    ServidorUnidade,
    TarefaInterna,
    TipoDemanda,
    TipoProducao,
    UnidadeInterna,
)


MSG_SEM_PERMISSAO = "Você não tem permissão para realizar esta ação."
MSG_OS_SOMENTE_LEITURA = (
    "Esta OS está em modo somente leitura para sua unidade. "
    "Solicite a reabertura à chefia da unidade responsável."
)

NIVEL_VISAO_SISTEMICA = "SISTEMICA"
NIVEL_VISAO_UNIDADE = "UNIDADE"
NIVEL_VISAO_PESSOAL = "PESSOAL"

PRIORIDADE_ORDEM = {
    "URGENTE": 0,
    "PRIORITARIO": 1,
    "NORMAL": 2,
}

PRIORIDADE_OS_LABELS = {
    "NORMAL": "Normal",
    "PRIORITARIO": "Prioritário",
    "URGENTE": "Urgente",
}

STATUS_PRODUCAO_FINAL = [
    Producao.STATUS_ENVIADO,
    Producao.STATUS_CANCELADO,
]

logger = logging.getLogger(__name__)


def _determinar_nivel_visao(perfil):
    if perfil and perfil.visibilidade_total:
        return NIVEL_VISAO_SISTEMICA
    if perfil and perfil.pode_homologar:
        return NIVEL_VISAO_UNIDADE
    return NIVEL_VISAO_PESSOAL


def _obter_unidade_principal_servidor(servidor):
    vinculo = obter_vinculo_unidade_ativo(servidor)
    return vinculo.unidade if vinculo else None


def _obter_visao_label(nivel_visao, servidor):
    if nivel_visao == NIVEL_VISAO_SISTEMICA:
        return "Visão sistêmica — Divisão de Avaliação de Imóveis"
    if nivel_visao == NIVEL_VISAO_UNIDADE:
        unidade = _obter_unidade_principal_servidor(servidor)
        if unidade:
            return f"Visão da unidade — {unidade.sigla}"
        return "Visão da unidade"
    return f"Minha visão — {servidor.nome}"


def _contexto_dashboard_vazio():
    return {
        "nivel_visao": NIVEL_VISAO_PESSOAL,
        "servidor_logado": None,
        "visao_label": "",
        "total_os_ativas": 0,
        "total_os_unidade": 0,
        "os_por_unidade": [],
        "os_prazo_proximo": [],
        "os_aguardando_retorno": [],
        "producao_por_tipo_mes": [],
        "producao_por_semana": [],
        "os_por_macroetapa": [],
        "os_por_natureza": [],
        "dashboard_chart_data": {},
        "card_aguard_retorno": 0,
        "card_producao_mes": 0,
        "card_prazo_proximo": 0,
    }


def _obter_servidor(user):
    try:
        return user.servidor
    except Servidor.DoesNotExist:
        return None


def _obter_unidades_ativas(servidor):
    hoje = timezone.localdate()
    return servidor.vinculos_unidade.filter(
        Q(data_fim__isnull=True) | Q(data_fim__gte=hoje),
    ).values_list("unidade_id", flat=True)


def _contar_os_abertas():
    return _contar_os_ativas()


def _contar_os_ativas(os_ids=None):
    qs = OS.objects.filter(encerrada=False)
    if os_ids is not None:
        if not os_ids:
            return 0
        qs = qs.filter(pk__in=os_ids)
    return qs.count()


def _obter_os_ids_unidades(unidades_ids):
    if not unidades_ids:
        return set()
    return set(
        TarefaInterna.objects.filter(unidade_id__in=unidades_ids)
        .exclude(status="CONCLUIDO")
        .values_list("os_id", flat=True)
        .distinct(),
    )


def _obter_os_prazo_proximo(os_ids=None):
    """OSs com entrada na divisão há mais de 25 dias e não encerradas."""
    hoje = timezone.localdate()
    corte = hoje - datetime.timedelta(days=25)

    ordens = (
        _queryset_os_anotado()
        .filter(prazo__lt=corte)
        .exclude(prazo__isnull=True)
        .filter(encerrada=False)
        .order_by("prazo")
    )
    if os_ids is not None:
        if not os_ids:
            return []
        ordens = ordens.filter(pk__in=os_ids)

    resultado = []
    for os_obj in ordens:
        dias = (hoje - os_obj.prazo).days
        dias_restantes = None
        if os_obj.prazo_data:
            dias_restantes = (os_obj.prazo_data - hoje).days
        resultado.append(
            {
                "pk": os_obj.pk,
                "numero_os": os_obj.numero_os,
                "processo_sei": os_obj.processo_sei_numero or "—",
                "dias": dias,
                "dias_restantes": dias_restantes,
                "prazo_data": os_obj.prazo_data,
            },
        )
    return sorted(resultado, key=lambda item: item["dias"], reverse=True)


def _obter_os_aguardando_retorno():
    encaminhamentos = (
        Encaminhamento.objects.filter(
            aguarda_retorno=True,
            data_retorno_efetiva__isnull=True,
        )
        .select_related("os", "unidade_externa_destino")
        .order_by("data_retorno_prevista", "os__numero_os")
    )
    return [
        {
            "os_pk": enc.os_id,
            "numero_os": enc.os.numero_os,
            "unidade_externa": (
                enc.unidade_externa_destino.nome
                if enc.unidade_externa_destino
                else "—"
            ),
            "data_retorno_prevista": enc.data_retorno_prevista,
        }
        for enc in encaminhamentos
    ]


def _obter_producao_por_tipo_mes(os_ids=None):
    hoje = timezone.localdate()
    queryset = Producao.objects.filter(
        status=Producao.STATUS_ENVIADO,
        data_enviado__year=hoje.year,
        data_enviado__month=hoje.month,
    ).select_related("tipo_producao")
    if os_ids is not None:
        if not os_ids:
            return []
        queryset = queryset.filter(os_id__in=os_ids)

    totais = {}
    for producao in queryset:
        label = (
            producao.tipo_producao.label_display
            if producao.tipo_producao
            else "—"
        )
        totais[label] = totais.get(label, 0) + 1

    return [
        {"prefixo": label, "total": total}
        for label, total in sorted(totais.items())
    ]


def _obter_producao_por_semana(os_ids=None):
    hoje = timezone.localdate()
    inicio = hoje - datetime.timedelta(days=7 * 8 - 1)
    resultado = []
    for indice in range(8):
        semana_inicio = inicio + datetime.timedelta(days=7 * indice)
        semana_fim = semana_inicio + datetime.timedelta(days=6)
        queryset = Producao.objects.filter(
            status=Producao.STATUS_ENVIADO,
            data_enviado__gte=semana_inicio,
            data_enviado__lte=semana_fim,
        )
        if os_ids is not None:
            if not os_ids:
                resultado.append({"semana": f"S{indice + 1}", "total": 0})
                continue
            queryset = queryset.filter(os_id__in=os_ids)
        resultado.append(
            {"semana": f"S{indice + 1}", "total": queryset.count()},
        )
    return resultado


def _contar_producao_homologada_mes(os_ids=None):
    hoje = timezone.localdate()
    queryset = Producao.objects.filter(
        status=Producao.STATUS_ENVIADO,
        data_enviado__year=hoje.year,
        data_enviado__month=hoje.month,
    )
    if os_ids is not None:
        if not os_ids:
            return 0
        queryset = queryset.filter(os_id__in=os_ids)
    return queryset.count()


def _obter_os_por_macroetapa(os_ids=None):
    qs = queryset_os_com_macroetapa().filter(encerrada=False)
    if os_ids is not None:
        if not os_ids:
            return []
        qs = qs.filter(pk__in=os_ids)

    linhas = (
        qs.values("macroetapa_atual")
        .annotate(total=Count("id"))
        .order_by("-total")
    )
    return [
        {"macroetapa": linha["macroetapa_atual"], "total": linha["total"]}
        for linha in linhas
    ]


def _obter_os_por_natureza(os_ids=None):
    qs = queryset_os_com_macroetapa().filter(encerrada=False)
    if os_ids is not None:
        if not os_ids:
            return []
        qs = qs.filter(pk__in=os_ids)

    linhas = (
        qs.values("natureza__descricao")
        .annotate(total=Count("id"))
        .order_by("-total")
    )
    return [
        {"natureza": linha["natureza__descricao"], "total": linha["total"]}
        for linha in linhas
    ]


def _contexto_dashboard_sistemica(servidor, unidades_ids, perfil):
    os_prazo_proximo = _obter_os_prazo_proximo()
    os_aguardando_retorno = _obter_os_aguardando_retorno()
    producao_por_tipo_mes = _obter_producao_por_tipo_mes()
    producao_por_semana = _obter_producao_por_semana()
    os_por_macroetapa = _obter_os_por_macroetapa()
    os_por_natureza = _obter_os_por_natureza()

    return {
        "total_os_ativas": _contar_os_ativas(),
        "os_por_unidade": os_ativas_por_unidade(),
        "os_prazo_proximo": os_prazo_proximo,
        "os_aguardando_retorno": os_aguardando_retorno,
        "producao_por_tipo_mes": producao_por_tipo_mes,
        "producao_por_semana": producao_por_semana,
        "os_por_macroetapa": os_por_macroetapa,
        "os_por_natureza": os_por_natureza,
        "dashboard_chart_data": _montar_dashboard_chart_data(
            producao_por_tipo_mes,
            producao_por_semana,
            os_por_macroetapa,
            os_por_natureza,
        ),
        "card_aguard_retorno": len(os_aguardando_retorno),
        "card_producao_mes": _contar_producao_homologada_mes(),
        "card_prazo_proximo": len(os_prazo_proximo),
    }


def _contexto_dashboard_unidade(servidor, unidades_ids, perfil):
    unidade = _obter_unidade_principal_servidor(servidor)
    os_ids = set(os_da_unidade_atual(unidade).values_list("pk", flat=True))
    os_prazo_proximo = _obter_os_prazo_proximo(os_ids=os_ids)
    producao_por_tipo_mes = _obter_producao_por_tipo_mes(os_ids=os_ids)
    producao_por_semana = _obter_producao_por_semana(os_ids=os_ids)

    return {
        "total_os_unidade": _contar_os_ativas(os_ids=os_ids),
        "os_prazo_proximo": os_prazo_proximo,
        "producao_por_tipo_mes": producao_por_tipo_mes,
        "producao_por_semana": producao_por_semana,
        "unidade_sigla": unidade.sigla if unidade else "",
        "dashboard_chart_data": _montar_dashboard_chart_data(
            producao_por_tipo_mes,
            producao_por_semana,
            [],
            [],
        ),
        "card_producao_mes": _contar_producao_homologada_mes(os_ids=os_ids),
    }


def _contexto_dashboard_pessoal(servidor, perfil):
    unidade = _obter_unidade_principal_servidor(servidor)
    os_ids = set(os_da_unidade_atual(unidade).values_list("pk", flat=True)) if unidade else set()
    return {
        "card_producao_mes": _contar_producao_homologada_mes(os_ids=os_ids),
    }


def _montar_dashboard_chart_data(
    producao_por_tipo_mes,
    producao_por_semana,
    os_por_macroetapa,
    os_por_natureza,
):
    from core.templatetags.siprac_filters import MACROETAPA_LABELS

    return {
        "producao_tipo": {
            "labels": [item["prefixo"] for item in producao_por_tipo_mes],
            "data": [item["total"] for item in producao_por_tipo_mes],
        },
        "producao_semana": {
            "labels": [item["semana"] for item in producao_por_semana],
            "data": [item["total"] for item in producao_por_semana],
        },
        "os_macroetapa": {
            "labels": [
                MACROETAPA_LABELS.get(item["macroetapa"], item["macroetapa"])
                for item in os_por_macroetapa
            ],
            "data": [item["total"] for item in os_por_macroetapa],
        },
        "os_natureza": {
            "labels": [item["natureza"] for item in os_por_natureza],
            "data": [item["total"] for item in os_por_natureza],
        },
    }


def _gerar_numero_os():
    ano = timezone.localdate().year
    sufixo = f"_{ano}"
    maior_sequencia = 0

    for numero in OS.objects.filter(numero_os__endswith=sufixo).values_list(
        "numero_os",
        flat=True,
    ):
        partes = numero.split("_")
        if len(partes) == 3 and partes[0] == "OS":
            try:
                maior_sequencia = max(maior_sequencia, int(partes[1]))
            except ValueError:
                continue

    return f"OS_{maior_sequencia + 1:05d}_{ano}"


def _obter_tipo_producao_despacho():
    tipo, _ = TipoProducao.objects.get_or_create(
        prefixo="Despacho",
        defaults={"descricao": "Despacho", "ativo": True},
    )
    return tipo


def _gerar_numero_producao(tipo_producao):
    ano = timezone.localdate().year
    prefixo = tipo_producao.prefixo
    maior_sequencia = 0

    for numero in Producao.objects.filter(
        tipo_producao=tipo_producao,
        ano=ano,
    ).exclude(numero_producao__isnull=True).exclude(numero_producao="").values_list(
        "numero_producao",
        flat=True,
    ):
        partes = numero.split("_")
        if len(partes) == 3 and partes[0] == prefixo and partes[2] == str(ano):
            try:
                maior_sequencia = max(maior_sequencia, int(partes[1]))
            except ValueError:
                continue

    return f"{prefixo}_{maior_sequencia + 1:03d}_{ano}"


def _gerar_codigo_isic():
    maior_sequencia = 0

    for codigo in Imovel.objects.exclude(codigo_isic__isnull=True).exclude(
        codigo_isic="",
    ).values_list("codigo_isic", flat=True):
        partes = codigo.split("_")
        if len(partes) == 2 and partes[0] == "ISIC":
            try:
                maior_sequencia = max(maior_sequencia, int(partes[1]))
            except ValueError:
                continue

    return f"ISIC_{maior_sequencia + 1:04d}"


def _ultimo_os_imovel(imovel):
    return imovel.os_imoveis.order_by("-data_vinculo", "-pk").first()


def _obter_os_imovel_vinculo(vinculo):
    if hasattr(vinculo, "os_imovel"):
        return vinculo.os_imovel
    return vinculo


def _salvar_imovel_from_form(dados, imovel=None):
    if imovel is None:
        imovel = Imovel(tipo_identificacao=dados["tipo_identificacao"])

    imovel.tipo_identificacao = dados["tipo_identificacao"]
    imovel.observacao_interna = dados.get("observacao_interna") or None

    if dados["tipo_identificacao"] == "ISIC":
        if not imovel.codigo_isic:
            imovel.codigo_isic = _gerar_codigo_isic()
        imovel.inscricao_cadastral = None
    else:
        imovel.inscricao_cadastral = dados["inscricao_cadastral"]
        imovel.codigo_isic = None

    imovel.save()
    return imovel


def _salvar_isic_from_form(dados):
    imovel = Imovel(
        tipo_identificacao="ISIC",
        codigo_isic=_gerar_codigo_isic(),
        observacao_interna=dados.get("observacao_interna") or None,
    )
    imovel.save()
    return imovel


def _decimal_para_json(valor):
    return str(valor) if valor is not None else None


def _format_area_mapa(area):
    if area is None:
        return "—"
    try:
        from decimal import Decimal

        value = Decimal(str(area))
        parts = f"{value:,.2f}".split(".")
        inteiro = parts[0].replace(",", ".")
        return f"{inteiro},{parts[1]}"
    except Exception:
        return str(area)


def _endereco_os_imovel(os_imovel):
    if os_imovel is None or not os_imovel.nom_logradouro:
        return "—"
    endereco = os_imovel.nom_logradouro
    if os_imovel.num_endereco:
        endereco = f"{endereco}, {os_imovel.num_endereco}"
    return endereco


def _endereco_imovel(imovel):
    return _endereco_os_imovel(_ultimo_os_imovel(imovel))


def _identificacao_imovel(imovel):
    if imovel.inscricao_cadastral:
        return str(imovel.inscricao_cadastral)
    if imovel.codigo_isic:
        return imovel.codigo_isic
    return "—"


def _identificacao_os_imovel(os_imovel):
    if os_imovel is None:
        return "—"
    if os_imovel.imovel.inscricao_cadastral:
        return str(os_imovel.imovel.inscricao_cadastral)
    if os_imovel.imovel.codigo_isic:
        return os_imovel.imovel.codigo_isic
    return "—"


def _coords_os_imovel(os_imovel):
    if os_imovel is None or os_imovel.latitude is None or os_imovel.longitude is None:
        return None, None
    return float(os_imovel.latitude), float(os_imovel.longitude)


def _montar_imoveis_coords_os(vinculos):
    com_coords = []
    sem_coords = 0
    for vinculo in vinculos:
        lat, lng = _coords_os_imovel(vinculo)
        if lat is None or lng is None:
            sem_coords += 1
            continue
        com_coords.append(
            {
                "id": vinculo.imovel_id,
                "identificacao": _identificacao_os_imovel(vinculo),
                "endereco": _endereco_os_imovel(vinculo),
                "area": _format_area_mapa(vinculo.area_territorial),
                "lat": lat,
                "lng": lng,
            },
        )
    return com_coords, sem_coords


def _imovel_para_mapa(imovel, os_imovel=None):
    os_imovel = os_imovel or _ultimo_os_imovel(imovel)
    lat, lng = _coords_os_imovel(os_imovel)
    return {
        "id": imovel.pk,
        "identificacao": _identificacao_imovel(imovel),
        "endereco": _endereco_os_imovel(os_imovel),
        "bairro": (os_imovel.bairro if os_imovel else None) or "—",
        "area": _format_area_mapa(
            os_imovel.area_territorial if os_imovel else None,
        ),
        "tipo_identificacao": imovel.tipo_identificacao,
        "lat": lat,
        "lng": lng,
    }


def _os_imovel_para_dict(os_imovel):
    if os_imovel is None:
        return {}
    return {
        "num_bloco": os_imovel.num_bloco,
        "cod_logradouro": os_imovel.cod_logradouro,
        "nom_logradouro": os_imovel.nom_logradouro,
        "num_endereco": os_imovel.num_endereco,
        "num_unidade": os_imovel.num_unidade,
        "bairro": os_imovel.bairro,
        "des_finalidade": os_imovel.des_finalidade,
        "area_territorial": os_imovel.area_territorial,
        "area_construida": os_imovel.area_construida,
        "exercicio_referencia": os_imovel.exercicio_referencia,
        "rh_nome": os_imovel.rh_nome,
        "rh_valor": os_imovel.rh_valor,
        "idf_regiao_homogenea": os_imovel.idf_regiao_homogenea,
        "latitude": os_imovel.latitude,
        "longitude": os_imovel.longitude,
        "coord_x": os_imovel.coord_x,
        "coord_y": os_imovel.coord_y,
        "origem_dados": os_imovel.origem_dados,
        "codigo_isic": os_imovel.imovel.codigo_isic,
    }


def _montar_resultado_busca(dados, fonte):
    endereco_partes = []
    if dados.get("nom_logradouro"):
        endereco_partes.append(dados["nom_logradouro"])
    if dados.get("num_endereco"):
        endereco_partes.append(str(dados["num_endereco"]))

    identificacao = dados.get("inscricao_cadastral") or dados.get("codigo_isic") or ""
    exercicio = dados.get("exercicio_referencia") or dados.get("exercicio")

    return {
        "inscricao_cadastral": dados.get("inscricao_cadastral"),
        "identificacao": str(identificacao),
        "endereco": ", ".join(endereco_partes),
        "bairro": dados.get("bairro") or "",
        "area_territorial": _format_area_mapa(dados.get("area_territorial")),
        "exercicio": exercicio,
        "num_versao": dados.get("num_versao", 0),
        "fonte": fonte,
        "num_bloco": dados.get("num_bloco") or "",
        "dados_completos": dados,
    }


def _anotar_prioridade_os(queryset):
    return queryset.annotate(
        prioridade_ordem=Case(
            When(prioridade="URGENTE", then=3),
            When(prioridade="PRIORITARIO", then=2),
            When(prioridade="NORMAL", then=1),
            default=0,
            output_field=IntegerField(),
        ),
    )


def _ordenar_queryset_os_fila(queryset):
    return _anotar_prioridade_os(queryset).order_by(
        "-prioridade_ordem",
        "-data_criacao_sgbd",
    )


def _queryset_os_anotado():
    processo_principal = OsProcesso.objects.filter(
        os_id=OuterRef("pk"),
        tipo_vinculo="PRINCIPAL",
    )

    base = OS.objects.select_related("natureza")
    return queryset_os_com_macroetapa(base).annotate(
        processo_sei_numero=Subquery(
            processo_principal.values("processo_sei__numero_processo")[:1],
        ),
        prazo=Subquery(processo_principal.values("data_entrada_divisao")[:1]),
    )


def _aplicar_visibilidade_os(queryset, request):
    perfil = getattr(request, "perfil_acesso", None)
    if perfil and perfil.visibilidade_total:
        return queryset

    servidor = _obter_servidor(request.user)
    if servidor is None:
        return queryset.none()

    unidades_ids = list(_obter_unidades_ativas(servidor))
    if not unidades_ids:
        return queryset.none()

    return queryset.filter(
        encaminhamentos__unidade_interna_destino_id__in=unidades_ids,
    ).distinct()


def _queryset_os_por_visibilidade(request):
    servidor = _obter_servidor(request.user)
    if not servidor:
        return OS.objects.none()

    visibilidade = getattr(request, "visibilidade", "UNIDADE")

    if visibilidade in ("TOTAL", "DEPARTAMENTO"):
        return OS.objects.filter(encerrada=False)

    # UNIDADE
    vinculo = getattr(request, "vinculo_ativo", None)
    if not vinculo:
        return OS.objects.none()
    return OS.objects.filter(
        encerrada=False,
        status_unidades__unidade=vinculo.unidade,
        status_unidades__status__in=(
            "ABERTA",
            "REABERTA",
            "CONCLUIDA",
        ),
    ).distinct()


def _aplicar_filtros_os(queryset, request):
    macroetapa = request.GET.get("macroetapa", "").strip()
    natureza_id = request.GET.get("natureza", "").strip()
    prioridade = request.GET.get("prioridade", "").strip()
    busca_processo = request.GET.get("q", "").strip()

    if macroetapa == "ativas":
        queryset = queryset.filter(encerrada=False)
    elif macroetapa == "ENCERRADO":
        queryset = queryset.filter(encerrada=True)
    elif macroetapa:
        queryset = queryset.filter(macroetapa_atual=macroetapa)
    if natureza_id:
        queryset = queryset.filter(natureza_id=natureza_id)
    if prioridade:
        queryset = queryset.filter(prioridade=prioridade)
    if busca_processo:
        queryset = queryset.filter(
            processos_vinculados__processo_sei__numero_processo__icontains=busca_processo,
        ).distinct()

    status_producao = request.GET.get("status_producao", "").strip()
    if status_producao:
        queryset = queryset.filter(producoes__status=status_producao).distinct()

    if request.GET.get("aguarda_retorno") == "1":
        os_ids = Encaminhamento.objects.filter(
            aguarda_retorno=True,
            data_retorno_efetiva__isnull=True,
        ).values_list("os_id", flat=True)
        queryset = queryset.filter(pk__in=os_ids)

    if request.GET.get("prazo") == "proximo":
        hoje = timezone.localdate()
        corte = hoje - datetime.timedelta(days=25)
        queryset = (
            queryset.filter(prazo__lt=corte)
            .exclude(prazo__isnull=True)
            .filter(encerrada=False)
        )

    unidade_sigla = request.GET.get("unidade", "").strip()
    if unidade_sigla:
        try:
            unidade = UnidadeInterna.objects.get(sigla=unidade_sigla)
            os_ids = os_da_unidade_atual(unidade).values_list("pk", flat=True)
            queryset = queryset.filter(pk__in=os_ids)
        except UnidadeInterna.DoesNotExist:
            pass

    return queryset


def _aplicar_visibilidade_producao(queryset, request):
    perfil = getattr(request, "perfil_acesso", None)
    servidor = _obter_servidor(request.user)
    if servidor is None:
        return queryset.none()

    if perfil and perfil.visibilidade_total:
        return queryset

    unidade = _obter_unidade_principal_servidor(servidor)
    if unidade is None:
        return queryset.none()
    os_ids = os_da_unidade_atual(unidade).values_list("pk", flat=True)
    return queryset.filter(os_id__in=os_ids)


def _aplicar_filtros_producao(queryset, request):
    status = request.GET.get("status", "").strip()
    if status:
        queryset = queryset.filter(status=status)

    tipo_id = request.GET.get("tipo_producao", "").strip()
    if tipo_id:
        queryset = queryset.filter(tipo_producao_id=tipo_id)

    periodo = request.GET.get("periodo", "").strip()
    if periodo == "mes_atual":
        hoje = timezone.localdate()
        queryset = queryset.filter(
            data_enviado__year=hoje.year,
            data_enviado__month=hoje.month,
            status=Producao.STATUS_ENVIADO,
        )

    unidade_sigla = request.GET.get("unidade", "").strip()
    if unidade_sigla:
        try:
            unidade = UnidadeInterna.objects.get(sigla=unidade_sigla)
            os_ids = os_da_unidade_atual(unidade).values_list("pk", flat=True)
            queryset = queryset.filter(os_id__in=os_ids)
        except UnidadeInterna.DoesNotExist:
            pass

    return queryset


def _obter_outras_os_mesmo_imovel(processo):
    os_ids_processo = set(
        OsProcesso.objects.filter(processo_sei=processo).values_list("os_id", flat=True),
    )
    if not os_ids_processo:
        return []

    imovel_ids = (
        OsImovel.objects.filter(os_id__in=os_ids_processo)
        .values_list("imovel_id", flat=True)
        .distinct()
    )
    if not imovel_ids:
        return []

    outros_os_ids = (
        OsImovel.objects.filter(imovel_id__in=imovel_ids)
        .exclude(os_id__in=os_ids_processo)
        .values_list("os_id", flat=True)
        .distinct()
    )
    if not outros_os_ids:
        return []

    vinculos_principais = {
        vinculo.os_id: vinculo
        for vinculo in OsProcesso.objects.filter(
            os_id__in=outros_os_ids,
            tipo_vinculo="PRINCIPAL",
        ).select_related("processo_sei", "os")
    }

    resultado = []
    vistos = set()
    for os_id in outros_os_ids:
        if os_id in vistos:
            continue
        vinculo = vinculos_principais.get(os_id)
        if vinculo is None or vinculo.processo_sei_id == processo.pk:
            continue
        vistos.add(os_id)
        resultado.append(
            {
                "processo_pk": vinculo.processo_sei_id,
                "numero_processo": vinculo.processo_sei.numero_processo,
                "os_pk": vinculo.os_id,
                "numero_os": vinculo.os.numero_os,
                "periodo_inicio": vinculo.data_entrada_divisao,
                "periodo_fim": vinculo.data_encerramento,
            },
        )
    return resultado


class SipracLoginView(LoginView):
    template_name = "login.html"
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse_lazy("dashboard")


class SipracLogoutView(LogoutView):
    next_page = reverse_lazy("login")


class TrocarPerfilView(RequerLoginMixin, View):
    def post(self, request):
        vinculo_id = request.POST.get("vinculo_id")
        servidor = _obter_servidor(request.user)

        if vinculo_id and servidor:
            try:
                hoje = timezone.localdate()
                vinculo = ServidorUnidade.objects.get(
                    pk=vinculo_id,
                    servidor=servidor,
                    data_inicio__lte=hoje,
                )
                if vinculo.data_fim is not None and vinculo.data_fim < hoje:
                    raise ServidorUnidade.DoesNotExist
                request.session["vinculo_ativo_id"] = vinculo.pk
            except (ServidorUnidade.DoesNotExist, ValueError, TypeError):
                pass

        if request.POST.get("limpar_wizard"):
            for key in (
                "wizard_inscricoes",
                "wizard_dados_os",
                "wizard_relacionado_os_pk",
            ):
                request.session.pop(key, None)
            return redirect("os_list")

        next_url = request.META.get("HTTP_REFERER") or "/"
        return redirect(next_url)


class DashboardView(RequerLoginMixin, TemplateView):
    template_name = "dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        perfil = getattr(self.request, "perfil_acesso", None)
        servidor = _obter_servidor(self.request.user)

        if servidor is None:
            context.update(_contexto_dashboard_vazio())
            return context

        nivel_visao = _determinar_nivel_visao(perfil)
        unidades_ids = list(_obter_unidades_ativas(servidor))

        context["nivel_visao"] = nivel_visao
        context["servidor_logado"] = servidor

        if nivel_visao == NIVEL_VISAO_SISTEMICA:
            context.update(_contexto_dashboard_sistemica(servidor, unidades_ids, perfil))
        elif nivel_visao == NIVEL_VISAO_UNIDADE:
            context.update(_contexto_dashboard_unidade(servidor, unidades_ids, perfil))
        else:
            context.update(_contexto_dashboard_pessoal(servidor, perfil))

        context["visao_label"] = _obter_visao_label(nivel_visao, servidor)
        return context


RELATORIOS_DISPONIVEIS = [
    {
        "titulo": "Produção por Servidor",
        "descricao": (
            "Produções homologadas agrupadas por autor do trabalho, com filtros "
            "por período, tipo de produção e unidade."
        ),
        "url_name": "relatorio_producao",
    },
    {
        "titulo": "Tempo de Registro de Processos",
        "descricao": (
            "Desempenho do auxiliar administrativo: intervalo entre a entrada "
            "do processo na Divisão e o registro no SIPRAC."
        ),
        "url_name": "relatorio_tempo_registro",
    },
]


def _filtros_relatorio_tempo_registro(form):
    dados = form.cleaned_data
    filtros = {}
    if dados.get("servidor"):
        filtros["servidor_id"] = dados["servidor"].pk
    if dados.get("data_inicio"):
        filtros["data_inicio"] = dados["data_inicio"]
    if dados.get("data_fim"):
        filtros["data_fim"] = dados["data_fim"]
    return filtros


def _filtros_relatorio_producao(form):
    dados = form.cleaned_data
    filtros = {}
    if dados.get("servidor"):
        filtros["servidor_id"] = dados["servidor"].pk
    if dados.get("data_inicio"):
        filtros["data_inicio"] = dados["data_inicio"]
    if dados.get("data_fim"):
        filtros["data_fim"] = dados["data_fim"]
    if dados.get("tipo_producao"):
        filtros["tipo_producao_id"] = dados["tipo_producao"].pk
    if dados.get("unidade"):
        filtros["unidade_id"] = dados["unidade"].pk
    return filtros


class RelatorioListView(RequerLoginMixin, TemplateView):
    template_name = "relatorio_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["relatorios"] = RELATORIOS_DISPONIVEIS
        return context


class RelatorioProducaoView(RequerLoginMixin, View):
    template_name = "relatorio_producao.html"

    def _render(self, request, form, linhas=None, total=None):
        return render(
            request,
            self.template_name,
            {
                "form": form,
                "linhas": linhas,
                "total": total,
                "exibir_resultados": linhas is not None,
            },
        )

    def get(self, request):
        return self._render(request, RelatorioProducaoForm())

    def post(self, request):
        form = RelatorioProducaoForm(request.POST)
        if not form.is_valid():
            return self._render(request, form)

        queryset = relatorio_producao_por_servidor(_filtros_relatorio_producao(form))
        acao = request.POST.get("action", "visualizar")

        if acao == "exportar":
            return exportar_producao_excel(queryset)

        linhas = linhas_relatorio_producao(queryset)
        return self._render(request, form, linhas=linhas, total=len(linhas))


class RelatorioTempoRegistroView(RequerLoginMixin, View):
    template_name = "relatorio_tempo_registro.html"

    def _render(self, request, form, linhas=None, total=None):
        return render(
            request,
            self.template_name,
            {
                "form": form,
                "linhas": linhas,
                "total": total,
                "exibir_resultados": linhas is not None,
            },
        )

    def get(self, request):
        return self._render(request, RelatorioTempoRegistroForm())

    def post(self, request):
        form = RelatorioTempoRegistroForm(request.POST)
        if not form.is_valid():
            return self._render(request, form)

        queryset = relatorio_tempo_registro_processo(
            _filtros_relatorio_tempo_registro(form),
        )
        acao = request.POST.get("action", "visualizar")

        if acao == "exportar":
            return exportar_tempo_registro_excel(queryset)

        linhas = linhas_relatorio_tempo_registro(queryset)
        return self._render(request, form, linhas=linhas, total=len(linhas))


class OSCreateView(RequerLoginMixin, FormView):
    """Mantida por compatibilidade; a criação oficial usa o wizard."""

    template_name = "os_form.html"
    form_class = OSForm

    def dispatch(self, request, *args, **kwargs):
        return redirect("os_nova")


def _wizard_requer_departamento(request):
    visibilidade = getattr(request, "visibilidade", "UNIDADE")
    return visibilidade in ("DEPARTAMENTO", "TOTAL")


def _wizard_json_safe(obj):
    if isinstance(obj, dict):
        return {chave: _wizard_json_safe(valor) for chave, valor in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_wizard_json_safe(valor) for valor in obj]
    if isinstance(obj, (datetime.date, datetime.datetime)):
        return obj.isoformat()
    if isinstance(obj, Decimal):
        return str(obj)
    return obj


def _buscar_dados_siat_inscricao(inscricao):
    try:
        inscricao_int = int(str(inscricao).strip())
    except (TypeError, ValueError):
        return None
    # Consulta O(1) no índice em memória
    dados = siat_index.buscar_por_inscricao(inscricao_int)
    if dados:
        return dados
    # Fallback para streaming só se o índice ainda não estiver pronto
    if not siat_index.indice_pronto():
        return buscar_inscricao_no_arquivo(inscricao_int, SIAT_ARQUIVO_PATH)
    return None


def _os_ativas_para_inscricao(inscricao):
    try:
        inscricao_int = int(str(inscricao).strip())
    except (TypeError, ValueError):
        return []

    processos_principais = Prefetch(
        "processos_vinculados",
        queryset=(
            OsProcesso.objects.filter(tipo_vinculo="PRINCIPAL")
            .select_related("processo_sei")
        ),
        to_attr="_processos_principais",
    )

    ordens = (
        OS.objects.filter(
            encerrada=False,
            os_imoveis__imovel__inscricao_cadastral=inscricao_int,
        )
        .prefetch_related(processos_principais)
        .distinct()
        .order_by("-data_criacao_sgbd")[:5]
    )

    resultado = []
    for os_obj in ordens:
        processo = None
        principais = getattr(os_obj, "_processos_principais", None) or []
        if principais:
            processo = principais[0]
        resultado.append(
            {
                "pk": os_obj.pk,
                "numero_os": os_obj.numero_os,
                "processo": (
                    processo.processo_sei.numero_processo if processo else ""
                ),
            },
        )
    return resultado


def _montar_resultado_wizard_inscricao(inscricao):
    inscricao_str = str(inscricao).strip()
    dados = _buscar_dados_siat_inscricao(inscricao_str)
    os_ativas = _os_ativas_para_inscricao(inscricao_str)

    if not dados:
        return {
            "inscricao": inscricao_str,
            "encontrada": False,
            "endereco": "",
            "bairro": "",
            "os_ativas": os_ativas,
            "dados_completos": None,
        }

    montado = _montar_resultado_busca(dados, "siat")
    return {
        "inscricao": str(dados.get("inscricao_cadastral") or inscricao_str),
        "encontrada": True,
        "endereco": montado["endereco"],
        "bairro": montado["bairro"] or "",
        "os_ativas": os_ativas,
        "dados_completos": _wizard_json_safe(dados),
    }


def _resolver_resultados_wizard_inscricoes(request, inscricoes):
    """Reutiliza resultados já na sessão para evitar reconsulta SIAT/DB."""
    cache = {
        str(item.get("inscricao")): item
        for item in (request.session.get("wizard_inscricoes") or [])
        if item.get("inscricao") is not None
    }
    resultados = []
    for inscricao in inscricoes:
        chave = str(inscricao).strip()
        if not chave:
            continue
        if chave in cache:
            resultados.append(cache[chave])
        else:
            resultados.append(_montar_resultado_wizard_inscricao(chave))
    return resultados


def _limpar_sessao_wizard(request):
    for chave in (
        "wizard_inscricoes",
        "wizard_dados_os",
        "wizard_relacionado_os_pk",
        "wizard_relacionado_pk",
    ):
        request.session.pop(chave, None)


class OSWizardCancelarView(RequerLoginMixin, View):
    def post(self, request, *args, **kwargs):
        _limpar_sessao_wizard(request)
        messages.info(request, "Criação de OS cancelada.")
        return redirect("os_list")


class OSWizardView(RequerLoginMixin, TemplateView):
    template_name = "os_wizard_passo1.html"

    def dispatch(self, request, *args, **kwargs):
        visibilidade = getattr(request, "visibilidade", "UNIDADE")
        if visibilidade not in ("DEPARTAMENTO", "TOTAL"):
            messages.error(
                request,
                "A criação de OS pelo wizard é restrita ao perfil DEPARTAMENTO.",
            )
            return redirect("os_list")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["wizard_inscricoes"] = self.request.session.get(
            "wizard_inscricoes",
            [],
        )
        context["passo"] = 1
        return context

    def post(self, request, *args, **kwargs):
        inscricoes = request.POST.getlist("inscricoes[]") or request.POST.getlist(
            "inscricoes",
        )
        if not inscricoes and request.content_type and "json" in request.content_type:
            try:
                payload = json.loads(request.body.decode("utf-8") or "{}")
            except (TypeError, ValueError, UnicodeDecodeError):
                payload = {}
            inscricoes = payload.get("inscricoes") or []

        resultados = _resolver_resultados_wizard_inscricoes(request, inscricoes)

        request.session["wizard_inscricoes"] = resultados
        request.session.modified = True

        payload = {
            "inscricoes_encontradas": resultados,
            "ha_os_ativa": any(item.get("os_ativas") for item in resultados),
        }

        wants_json = (
            request.headers.get("X-Requested-With") == "XMLHttpRequest"
            or "application/json" in (request.headers.get("Accept") or "")
            or request.GET.get("format") == "json"
        )
        if wants_json:
            return JsonResponse(payload)

        if not resultados:
            messages.error(request, "Informe ao menos uma inscrição cadastral.")
            return redirect("os_nova")

        return redirect("os_nova_passo2")


class OSUploadInscricoesView(RequerLoginMixin, View):
    def dispatch(self, request, *args, **kwargs):
        if not _wizard_requer_departamento(request):
            return JsonResponse({"error": "Sem permissão."}, status=403)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        arquivo = request.FILES.get("arquivo") or request.FILES.get("file")
        if not arquivo:
            return JsonResponse({"error": "Arquivo não enviado."}, status=400)
        if not arquivo.name.lower().endswith((".xlsx", ".xlsm")):
            return JsonResponse(
                {"error": "Envie um arquivo .xlsx."},
                status=400,
            )

        try:
            from openpyxl import load_workbook

            workbook = load_workbook(arquivo, read_only=True, data_only=True)
            sheet = workbook.active
            inscricoes = []
            for idx, row in enumerate(sheet.iter_rows(
                min_col=1, max_col=1, values_only=True
            )):
                valor = row[0]
                if valor is None:
                    continue
                texto = str(valor).strip()
                if not texto:
                    continue
                # Ignorar cabeçalho (primeira linha sem dígitos)
                if idx == 0 and not texto.isdigit():
                    continue
                # Aceitar apenas células com dígitos puros
                if not texto.isdigit():
                    continue
                inscricoes.append(texto)
            workbook.close()
        except Exception as exc:  # noqa: BLE001
            return JsonResponse(
                {"error": f"Falha ao ler planilha: {exc}"},
                status=400,
            )

        # remove duplicadas preservando ordem
        vistos = set()
        unicos = []
        for insc in inscricoes:
            if insc in vistos:
                continue
            vistos.add(insc)
            unicos.append(insc)

        resultados = [_montar_resultado_wizard_inscricao(insc) for insc in unicos]
        request.session["wizard_inscricoes"] = resultados
        request.session.modified = True
        return JsonResponse(
            {
                "inscricoes_encontradas": resultados,
                "ha_os_ativa": any(item.get("os_ativas") for item in resultados),
            },
        )


class OSWizardPasso2View(RequerLoginMixin, FormView):
    template_name = "os_wizard_passo2.html"

    def dispatch(self, request, *args, **kwargs):
        if not _wizard_requer_departamento(request):
            messages.error(
                request,
                "A criação de OS pelo wizard é restrita ao perfil DEPARTAMENTO.",
            )
            return redirect("os_list")
        if not request.session.get("wizard_inscricoes"):
            messages.error(request, "Adicione ao menos uma inscrição no passo 1.")
            return redirect("os_nova")

        relacionado = request.GET.get("relacionado") or request.session.get(
            "wizard_relacionado_os_pk",
        )
        self.os_relacionada = None
        if relacionado:
            self.os_relacionada = get_object_or_404(OS, pk=relacionado)
            request.session["wizard_relacionado_os_pk"] = self.os_relacionada.pk
        else:
            request.session.pop("wizard_relacionado_os_pk", None)

        return super().dispatch(request, *args, **kwargs)

    def get_form_class(self):
        if self.os_relacionada:
            return OSWizardPasso2RelacionadoForm
        return OSWizardPasso2Form

    def get_initial(self):
        initial = super().get_initial()
        dados = self.request.session.get("wizard_dados_os") or {}
        for campo in (
            "numero_processo",
            "data_abertura_sei",
            "data_entrada_divisao",
            "prioridade",
            "observacao",
            "apelido",
            "prazo_tipo",
            "prazo_data",
        ):
            if dados.get(campo) not in (None, ""):
                initial[campo] = dados[campo]
        if not self.os_relacionada:
            for campo in ("natureza", "tipo_demanda", "finalidade"):
                if dados.get(campo):
                    initial[campo] = dados[campo]
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["passo"] = 2
        context["wizard_inscricoes"] = self.request.session.get(
            "wizard_inscricoes",
            [],
        )
        context["os_relacionada"] = self.os_relacionada
        context["modo_relacionado"] = bool(self.os_relacionada)
        return context

    def form_valid(self, form):
        cleaned = form.cleaned_data
        dados = {
            "numero_processo": cleaned["numero_processo"],
            "data_abertura_sei": cleaned["data_abertura_sei"].isoformat(),
            "data_entrada_divisao": cleaned["data_entrada_divisao"].isoformat(),
        }
        if self.os_relacionada:
            dados["relacionado_os_pk"] = self.os_relacionada.pk
        else:
            dados.update(
                {
                    "natureza": cleaned["natureza"].pk,
                    "tipo_demanda": cleaned["tipo_demanda"].pk,
                    "finalidade": cleaned["finalidade"].pk,
                    "prioridade": cleaned["prioridade"],
                    "observacao": cleaned.get("observacao") or "",
                    "apelido": cleaned.get("apelido") or "",
                    "prazo_tipo": cleaned.get("prazo_tipo") or "SEM_PRIORIDADE",
                    "prazo_data": (
                        cleaned["prazo_data"].isoformat()
                        if cleaned.get("prazo_data")
                        else None
                    ),
                },
            )
        self.request.session["wizard_dados_os"] = dados
        self.request.session.modified = True
        return redirect("os_nova_passo3")


class OSWizardPasso3View(RequerLoginMixin, FormView):
    template_name = "os_wizard_passo3.html"
    form_class = OSWizardPasso3Form

    def dispatch(self, request, *args, **kwargs):
        if not _wizard_requer_departamento(request):
            messages.error(
                request,
                "A criação de OS pelo wizard é restrita ao perfil DEPARTAMENTO.",
            )
            return redirect("os_list")
        if not request.session.get("wizard_inscricoes"):
            messages.error(request, "Adicione ao menos uma inscrição no passo 1.")
            return redirect("os_nova")
        if not request.session.get("wizard_dados_os"):
            messages.error(request, "Preencha os dados da OS no passo 2.")
            return redirect("os_nova_passo2")
        return super().dispatch(request, *args, **kwargs)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["unidade_interna_destino"].queryset = (
            UnidadeInterna.objects.filter(tipo="OPERACIONAL").order_by("sigla")
        )
        form.fields["unidade_interna_destino"].required = False
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dados = self.request.session.get("wizard_dados_os") or {}
        context["passo"] = 3
        context["wizard_inscricoes"] = self.request.session.get(
            "wizard_inscricoes",
            [],
        )
        context["wizard_dados_os"] = dados
        relacionado_pk = dados.get("relacionado_os_pk") or self.request.session.get(
            "wizard_relacionado_os_pk",
        )
        context["os_relacionada"] = (
            OS.objects.filter(pk=relacionado_pk).first() if relacionado_pk else None
        )
        context["modo_relacionado"] = bool(context["os_relacionada"])
        if not context["modo_relacionado"]:
            context["natureza"] = Natureza.objects.filter(
                pk=dados.get("natureza"),
            ).first()
            context["tipo_demanda"] = TipoDemanda.objects.filter(
                pk=dados.get("tipo_demanda"),
            ).first()
            context["finalidade"] = Finalidade.objects.filter(
                pk=dados.get("finalidade"),
            ).first()
        return context

    def post(self, request, *args, **kwargs):
        encaminhar = request.POST.get("encaminhar") == "True"
        if not encaminhar:
            return self._concluir(request, encaminhar=False)

        form = self.get_form()
        if form.is_valid():
            if not form.cleaned_data.get("unidade_interna_destino"):
                form.add_error(
                    "unidade_interna_destino",
                    "Selecione a unidade destino para encaminhar.",
                )
                return self.form_invalid(form)
            return self._concluir(
                request,
                encaminhar=True,
                dados_encaminhamento=form.cleaned_data,
            )
        return self.form_invalid(form)

    def _concluir(self, request, encaminhar, dados_encaminhamento=None):
        servidor = _obter_servidor(request.user)
        if servidor is None:
            messages.error(request, MSG_SEM_PERMISSAO)
            return redirect("os_list")

        dados = request.session.get("wizard_dados_os") or {}
        inscricoes = request.session.get("wizard_inscricoes") or []
        relacionado_pk = dados.get("relacionado_os_pk") or request.session.get(
            "wizard_relacionado_os_pk",
        )

        try:
            data_abertura = datetime.date.fromisoformat(dados["data_abertura_sei"])
            data_entrada = datetime.date.fromisoformat(dados["data_entrada_divisao"])
        except (KeyError, TypeError, ValueError):
            messages.error(request, "Dados do passo 2 incompletos.")
            return redirect("os_nova_passo2")

        with transaction.atomic():
            processo_sei, _ = ProcessoSei.objects.get_or_create(
                numero_processo=dados["numero_processo"],
            )
            if not processo_sei.data_abertura_sei:
                processo_sei.data_abertura_sei = data_abertura
                processo_sei.save(update_fields=["data_abertura_sei"])

            if relacionado_pk:
                os_obj = get_object_or_404(OS, pk=relacionado_pk)
                OsProcesso.objects.create(
                    os=os_obj,
                    processo_sei=processo_sei,
                    tipo_vinculo="RELACIONADO",
                    data_entrada_divisao=data_entrada,
                    registrado_por=servidor,
                )
            else:
                natureza = get_object_or_404(Natureza, pk=dados["natureza"])
                tipo_demanda = get_object_or_404(TipoDemanda, pk=dados["tipo_demanda"])
                finalidade = get_object_or_404(Finalidade, pk=dados["finalidade"])
                prazo_data = None
                if dados.get("prazo_data"):
                    prazo_data = datetime.date.fromisoformat(dados["prazo_data"])

                os_obj = OS.objects.create(
                    numero_os=_gerar_numero_os(),
                    data_entrada_divisao=data_entrada,
                    natureza=natureza,
                    tipo_demanda=tipo_demanda,
                    finalidade=finalidade,
                    prioridade=dados.get("prioridade") or "NORMAL",
                    observacao=dados.get("observacao") or None,
                    apelido=dados.get("apelido") or None,
                    prazo_tipo=dados.get("prazo_tipo") or "SEM_PRIORIDADE",
                    prazo_data=prazo_data,
                    criado_por=servidor,
                    pendente_encaminhamento=not encaminhar,
                )
                OsProcesso.objects.create(
                    os=os_obj,
                    processo_sei=processo_sei,
                    tipo_vinculo="PRINCIPAL",
                    data_entrada_divisao=data_entrada,
                    registrado_por=servidor,
                )

            for item in inscricoes:
                dados_siat = item.get("dados_completos")
                if not dados_siat:
                    continue
                inscricao = dados_siat.get("inscricao_cadastral")
                if inscricao is None:
                    continue
                imovel = Imovel.objects.filter(
                    inscricao_cadastral=inscricao,
                ).first()
                if imovel and OsImovel.objects.filter(os=os_obj, imovel=imovel).exists():
                    continue
                vincular_imovel_a_os(os_obj, dados_siat, servidor=servidor)

            if encaminhar and dados_encaminhamento:
                unidade_destino = dados_encaminhamento["unidade_interna_destino"]
                servidor_destino = dados_encaminhamento.get("servidor_destino")
                agora = timezone.now()
                encaminhamento = Encaminhamento.objects.create(
                    os=os_obj,
                    unidade_interna_origem=None,
                    servidor_origem=servidor,
                    unidade_interna_destino=unidade_destino,
                    servidor_destino=servidor_destino,
                    etapa_interna="ENTRADA",
                    tipo_macroetapa=Encaminhamento.TIPO_MACROETAPA_ATENDIMENTO_INTERNO,
                    automatico=False,
                    aguarda_retorno=False,
                    observacao=dados_encaminhamento.get("observacao") or None,
                    manter_aberta_na_unidade=False,
                )
                TarefaInterna.objects.create(
                    os=os_obj,
                    encaminhamento=encaminhamento,
                    unidade=unidade_destino,
                    servidor=servidor_destino or servidor,
                    etapa_interna="ENTRADA",
                    status="PENDENTE",
                    data_inicio=agora,
                )
                _atualizar_status_unidade_encaminhamento(
                    os_obj,
                    None,
                    servidor,
                    False,
                    unidade_destino=unidade_destino,
                )
                if os_obj.pendente_encaminhamento:
                    os_obj.pendente_encaminhamento = False
                    os_obj.save(update_fields=["pendente_encaminhamento"])

        _limpar_sessao_wizard(request)

        if encaminhar:
            messages.success(
                request,
                f"{os_obj.numero_os} criada e encaminhada com sucesso.",
            )
        elif relacionado_pk:
            messages.success(
                request,
                f"Processo relacionado registrado em {os_obj.numero_os}.",
            )
        else:
            messages.warning(
                request,
                f"{os_obj.numero_os} salva sem encaminhamento. "
                "Encaminhe para uma unidade operacional quando estiver pronto.",
            )
        return redirect(reverse("os_detalhe", kwargs={"pk": os_obj.pk}))


class ProcessoSeiAPIView(RequerLoginJSONMixin, View):
    def get(self, request, *args, **kwargs):
        numero = (request.GET.get("numero") or "").strip()
        if not numero:
            return JsonResponse(
                {
                    "encontrado": False,
                    "data_abertura_sei": None,
                    "os_ativas": [],
                },
            )

        processo = ProcessoSei.objects.filter(numero_processo=numero).first()
        if not processo:
            return JsonResponse(
                {
                    "encontrado": False,
                    "data_abertura_sei": None,
                    "os_ativas": [],
                },
            )

        os_ativas = []
        vinculos = (
            OsProcesso.objects.filter(
                processo_sei=processo,
                os__encerrada=False,
            )
            .select_related("os")
            .order_by("-os__data_criacao_sgbd")
        )
        for vinculo in vinculos:
            os_ativas.append(
                {
                    "pk": vinculo.os.pk,
                    "numero_os": vinculo.os.numero_os,
                    "tipo_vinculo": vinculo.tipo_vinculo,
                },
            )

        return JsonResponse(
            {
                "encontrado": True,
                "data_abertura_sei": (
                    processo.data_abertura_sei.isoformat()
                    if processo.data_abertura_sei
                    else None
                ),
                "os_ativas": os_ativas,
            },
        )


COLUNAS_GERENCIAL_CONFIG = {
    "entrada_dai": {"label": "Entrada DAI"},
    "entrada_eav": {"label": "Entrada EAV"},
    "requerimento": {"label": "Requerimento"},
    "finalidade": {"label": "Finalidade"},
    "ctm": {"label": "CTM"},
    "logradouro": {"label": "Logradouro"},
    "num_endereco": {"label": "Nº"},
    "num_unidade": {"label": "Unidade"},
    "num_bloco": {"label": "Lote Fiscal"},
    "numero_imovel": {"label": "Nº Imóvel"},
    "finalidade_imovel": {"label": "Finalidade Imóvel"},
    "area_territorial": {"label": "Área Territorial"},
    "area_construida": {"label": "Área Construída"},
    "bairro": {"label": "Bairro"},
    "rh_valor": {"label": "RH_VALOR"},
    "apelido": {"label": "Apelido"},
    "prioridade": {"label": "PRIORIDADE"},
    "dias_sei": {"label": "DIAS_SEI"},
    "enviado": {"label": "ENVIADO"},
    "la_pt_ptf": {"label": "LA_PT_PTF"},
    "tipo_trabalho": {"label": "TIPO_TRABALHO"},
    "doc_sei": {"label": "DOC_SEI"},
    "destino": {"label": "DESTINO"},
}

GRUPOS_COLUNAS_GERENCIAL = [
    (
        "INFORMAÇÕES DE ENTRADA",
        ["entrada_dai", "entrada_eav", "requerimento", "finalidade"],
    ),
    (
        "INFORMAÇÕES DO IMÓVEL",
        [
            "ctm",
            "logradouro",
            "num_endereco",
            "num_unidade",
            "num_bloco",
            "numero_imovel",
            "finalidade_imovel",
            "area_territorial",
            "area_construida",
            "bairro",
            "rh_valor",
        ],
    ),
    (
        "TRIAGEM",
        [
            "apelido",
            "prioridade",
            "dias_sei",
        ],
    ),
    (
        "PRAZOS",
        [
            "enviado",
        ],
    ),
    (
        "PRODUTOS",
        ["la_pt_ptf", "tipo_trabalho", "doc_sei", "destino"],
    ),
]

COLUNAS_GERENCIAL_NOVAS = {
    "entrada_eav",
    "num_endereco",
    "num_unidade",
    "num_bloco",
    "finalidade_imovel",
    "area_territorial",
    "area_construida",
    "bairro",
    "rh_valor",
    "apelido",
    "prioridade",
    "enviado",
    "la_pt_ptf",
    "doc_sei",
    "destino",
}

COLUNAS_GERENCIAL_PADRAO = [
    "entrada_dai",
    "entrada_eav",
    "requerimento",
    "finalidade",
    "ctm",
    "logradouro",
    "num_endereco",
    "num_unidade",
    "num_bloco",
    "finalidade_imovel",
    "area_territorial",
    "area_construida",
    "rh_valor",
    "numero_imovel",
    "bairro",
    "apelido",
    "prioridade",
    "enviado",
    "tipo_trabalho",
    "dias_sei",
    "la_pt_ptf",
    "doc_sei",
    "destino",
]

STATUS_GERENCIAL_CARDS = [
    Producao.STATUS_ENVIADO,
    Producao.STATUS_CANCELADO,
]


def _query_string_os_list(request, **overrides):
    params = request.GET.copy()
    params.pop("page", None)
    for chave, valor in overrides.items():
        if valor is None:
            params.pop(chave, None)
        elif valor == "":
            params.pop(chave, None)
        else:
            params[chave] = valor
    return params.urlencode()


def _query_string_filtros_os_list(request):
    params = request.GET.copy()
    params.pop("page", None)
    params.pop("view", None)
    for coluna in COLUNAS_GERENCIAL_CONFIG:
        params.pop(f"fg_{coluna}", None)
    return params.urlencode()


def _query_string_gerencial(request, **overrides):
    params = request.GET.copy()
    params.pop("page", None)
    for coluna in COLUNAS_GERENCIAL_CONFIG:
        params.pop(f"fg_{coluna}", None)
    params["view"] = "gerencial"
    for chave, valor in overrides.items():
        if valor is None:
            params.pop(chave, None)
        elif valor == "":
            params.pop(chave, None)
        else:
            params[chave] = valor
    return params.urlencode()


def _colunas_visiveis_gerencial(servidor):
    if servidor is None:
        saved_set = set(COLUNAS_GERENCIAL_PADRAO)
    else:
        try:
            preferencia = servidor.preferencia_gerencial
            colunas = preferencia.colunas_visiveis or []
            if colunas:
                colunas_salvas = set(colunas)
                if not colunas_salvas.intersection(COLUNAS_GERENCIAL_NOVAS):
                    preferencia.colunas_visiveis = list(COLUNAS_GERENCIAL_PADRAO)
                    preferencia.save(update_fields=["colunas_visiveis"])
                    colunas = preferencia.colunas_visiveis
        except PreferenciaGerencial.DoesNotExist:
            colunas = list(COLUNAS_GERENCIAL_PADRAO)
        saved_set = set(colunas) if colunas else set(COLUNAS_GERENCIAL_PADRAO)
    ordenadas = [c for c in COLUNAS_GERENCIAL_CONFIG if c in saved_set]
    return ordenadas or list(COLUNAS_GERENCIAL_PADRAO)


def _grupos_header_gerencial(colunas_visiveis):
    visiveis = set(colunas_visiveis)
    grupos = []
    for titulo, colunas in GRUPOS_COLUNAS_GERENCIAL:
        count = sum(1 for coluna in colunas if coluna in visiveis)
        if count > 0:
            grupos.append({"titulo": titulo, "colspan": count})
    return grupos


def _pode_editar_entrada_dai(request):
    perfil = getattr(request, "perfil_acesso", None)
    if perfil is None:
        return False
    return (
        perfil.pode_criar_os
        or perfil.pode_homologar
        or perfil.visibilidade_total
    )


def _formatar_data_br(data):
    if not data:
        return "—"
    return data.strftime("%d/%m/%Y")


def _formatar_decimal_br(valor):
    if valor is None:
        return "—"
    try:
        from decimal import Decimal

        value = Decimal(str(valor))
        parts = f"{value:,.2f}".split(".")
        inteiro = parts[0].replace(",", ".")
        return f"{inteiro},{parts[1]}"
    except Exception:
        return str(valor) if valor else "—"


def _destino_pos_homologacao(os_obj, producao):
    if not producao or not producao.data_enviado:
        return "—"
    enc = (
        Encaminhamento.objects.filter(
            os=os_obj,
            data_hora__date__gte=producao.data_enviado,
        )
        .select_related("unidade_externa_destino", "unidade_interna_destino")
        .order_by("-data_hora")
        .first()
    )
    if not enc:
        return "—"
    if enc.unidade_externa_destino:
        return enc.unidade_externa_destino.nome
    if enc.unidade_interna_destino:
        return enc.unidade_interna_destino.sigla
    return "—"


def _label_macroetapa_gerencial(macroetapa):
    return {
        "ENTRADA_DIVISAO": "Entrada na Divisão",
        "ATENDIMENTO_INTERNO": "Atend. Interno",
        "ATENDIMENTO_EXTERNO": "Atend. Externo",
        "RETORNO_EXTERNO": "Retorno Externo",
        "INCLUSAO_PROCESSO": "Inclusão de Processo",
        "ENCERRADO": "Encerrado",
        "ENCERRAMENTO": "Encerrado",
    }.get(macroetapa, macroetapa or "—")


def _label_status_producao_gerencial(status):
    return dict(Producao.STATUS_CHOICES).get(status, status or "—")


def _cor_status_producao_gerencial(status):
    return {
        "NAO_DISTRIBUIDO": "secondary",
        "DISTRIBUIDO": "info",
        "REVISAR": "warning",
        "REVISADO": "info",
        "VER_AJUSTES": "warning",
        "ENTREGA_AJUSTES": "warning",
        "AJUSTES_OK": "success",
        "HOMOLOGAR": "primary",
        "ENVIADO": "success",
        "CANCELADO": "danger",
    }.get(status, "secondary")


def _calcular_dias_sei_gerencial(os_obj):
    if not os_obj.prazo_data:
        return None
    return (os_obj.prazo_data - timezone.localdate()).days


def _formatar_numero_imovel_gerencial(oi):
    if not oi or not oi.imovel:
        return "—"
    if oi.imovel.inscricao_cadastral:
        return str(oi.imovel.inscricao_cadastral)
    return oi.imovel.codigo_isic or "—"


def _campos_imovel_vazios_gerencial():
    return {
        "ctm": "—",
        "logradouro": "—",
        "num_endereco": "—",
        "num_unidade": "—",
        "num_bloco": "—",
        "numero_imovel": "—",
        "numero_imovel_extra": 0,
        "numero_imovel_tooltip": "",
        "finalidade_imovel": "—",
        "area_territorial": "—",
        "area_construida": "—",
        "bairro": "—",
        "rh_valor": "—",
    }


def _campos_vazios_gerencial():
    return {
        "prioridade": "—",
        "dias_sei": None,
        "enviado": "—",
        "la_pt_ptf": "—",
        "tipo_trabalho": "—",
        "doc_sei": "—",
        "destino": "—",
        "producao_pk": None,
        "tem_producao": False,
    }


def _dados_imovel_gerencial(os_obj):
    """Dados do primeiro imóvel vinculado à OS."""
    oi = os_obj.os_imoveis.select_related("imovel").first()
    if not oi:
        return _campos_imovel_vazios_gerencial()
    return {
        "ctm": str(oi.cod_logradouro) if oi.cod_logradouro else "—",
        "logradouro": oi.nom_logradouro or "—",
        "num_endereco": oi.num_endereco or "—",
        "num_unidade": oi.num_unidade or "—",
        "num_bloco": oi.num_bloco or "—",
        "numero_imovel": _formatar_numero_imovel_gerencial(oi),
        "numero_imovel_extra": 0,
        "numero_imovel_tooltip": "",
        "finalidade_imovel": oi.des_finalidade or "—",
        "area_territorial": _formatar_decimal_br(oi.area_territorial),
        "area_construida": _formatar_decimal_br(oi.area_construida),
        "bairro": oi.bairro or "—",
        "rh_valor": str(oi.rh_valor) if oi.rh_valor else "—",
    }


def _dados_imoveis_producao_gerencial(producao):
    """Dados de imóveis da produção (primeiro completo + contador)."""
    imoveis = list(
        producao.producao_imoveis.select_related("os_imovel__imovel").all(),
    )
    total = len(imoveis)
    if total == 0:
        return _campos_imovel_vazios_gerencial()

    primeiro_oi = imoveis[0].os_imovel
    dados = {
        "ctm": str(primeiro_oi.cod_logradouro) if primeiro_oi.cod_logradouro else "—",
        "logradouro": primeiro_oi.nom_logradouro or "—",
        "num_endereco": primeiro_oi.num_endereco or "—",
        "num_unidade": primeiro_oi.num_unidade or "—",
        "num_bloco": primeiro_oi.num_bloco or "—",
        "finalidade_imovel": primeiro_oi.des_finalidade or "—",
        "area_territorial": _formatar_decimal_br(primeiro_oi.area_territorial),
        "area_construida": _formatar_decimal_br(primeiro_oi.area_construida),
        "bairro": primeiro_oi.bairro or "—",
        "rh_valor": str(primeiro_oi.rh_valor) if primeiro_oi.rh_valor else "—",
    }
    num_principal = _formatar_numero_imovel_gerencial(primeiro_oi)
    if total > 1:
        todos = [_formatar_numero_imovel_gerencial(pi.os_imovel) for pi in imoveis]
        dados["numero_imovel"] = num_principal
        dados["numero_imovel_extra"] = total - 1
        dados["numero_imovel_tooltip"] = ", ".join(todos)
    else:
        dados["numero_imovel"] = num_principal
        dados["numero_imovel_extra"] = 0
        dados["numero_imovel_tooltip"] = ""
    return dados


def _montar_cells_gerencial(linha):
    """Monta dict cells a partir dos campos flat da linha."""
    return {
        "entrada_dai": linha.get("entrada_dai", "—"),
        "entrada_eav": linha.get("entrada_eav", "—"),
        "requerimento": linha.get("requerimento", "—"),
        "finalidade": linha.get("finalidade", "—"),
        "ctm": linha.get("ctm", "—"),
        "logradouro": linha.get("logradouro", "—"),
        "num_endereco": linha.get("num_endereco", "—"),
        "num_unidade": linha.get("num_unidade", "—"),
        "num_bloco": linha.get("num_bloco", "—"),
        "numero_imovel": linha.get("numero_imovel", "—"),
        "finalidade_imovel": linha.get("finalidade_imovel", "—"),
        "area_territorial": linha.get("area_territorial", "—"),
        "area_construida": linha.get("area_construida", "—"),
        "bairro": linha.get("bairro", "—"),
        "rh_valor": linha.get("rh_valor", "—"),
        "apelido": linha.get("apelido") or "—",
        "prioridade": linha.get("prioridade", "—"),
        "dias_sei": linha.get("dias_sei"),
        "enviado": linha.get("enviado", "—"),
        "la_pt_ptf": linha.get("la_pt_ptf", "—"),
        "tipo_trabalho": linha.get("tipo_trabalho", "—"),
        "doc_sei": linha.get("doc_sei", "—"),
        "destino": linha.get("destino", "—"),
    }


def _cor_macroetapa_gerencial(macroetapa):
    return {
        "ENTRADA_DIVISAO": "secondary",
        "ATENDIMENTO_INTERNO": "primary",
        "ATENDIMENTO_EXTERNO": "info",
        "RETORNO_EXTERNO": "warning",
        "NOTIFICACAO": "dark",
        "ENCERRADO": "dark",
        "ENCERRAMENTO": "dark",
    }.get(macroetapa, "secondary")


def _serializar_producao_painel_gerencial(producao, request):
    logs = (
        ProducaoStatusLog.objects.filter(producao=producao)
        .select_related("servidor_origem", "servidor_destino")
        .order_by("-data_hora")[:10]
    )
    total_comentarios = Comentario.objects.filter(producao=producao).count()
    return {
        "pk": producao.pk,
        "label": (
            producao.tipo_producao.label_display
            if producao.tipo_producao
            else "—"
        ),
        "prefixo": (
            producao.tipo_producao.prefixo if producao.tipo_producao else "—"
        ),
        "status": producao.status,
        "status_label": _label_status_producao_gerencial(producao.status),
        "status_cor": _cor_status_producao_gerencial(producao.status),
        "pode_cancelar": producao.status != Producao.STATUS_CANCELADO,
        "enviado_iso": (
            producao.data_enviado.isoformat() if producao.data_enviado else ""
        ),
        "numero_producao": producao.numero_producao or "",
        "numero_sei": producao.numero_sei or "",
        "observacao": producao.observacao or "",
        "status_log": [_serializar_status_log(log) for log in logs],
        "total_comentarios": total_comentarios,
        "opcoes_pos_enviado": (
            _opcoes_pos_enviado(producao.os)
            if producao.status == Producao.STATUS_ENVIADO
            else []
        ),
    }


def _producoes_painel_gerencial(
    os_obj,
    unidade,
    servidor_logado,
    perfil_pode_homologar,
    request,
    producao_ativa=None,
    modo_b=False,
):
    if unidade:
        servidores_unidade_ids = ServidorUnidade.objects.filter(
            unidade=unidade,
            data_fim__isnull=True,
        ).values_list("servidor_id", flat=True)
        producoes = (
            os_obj.producoes.exclude(status=Producao.STATUS_CANCELADO)
            .filter(
                Q(unidade=unidade)
                | Q(criado_por_id__in=servidores_unidade_ids),
            )
            .select_related("tipo_producao")
            .order_by("-data_criacao")
        )
    else:
        producoes = (
            os_obj.producoes.exclude(status=Producao.STATUS_CANCELADO)
            .select_related("tipo_producao")
            .order_by("-data_criacao")
        )

    producoes = list(producoes)
    if not modo_b and producao_ativa is not None:
        producoes = [p for p in producoes if p.pk == producao_ativa.pk]

    return [
        _serializar_producao_painel_gerencial(producao, request)
        for producao in producoes
    ]


def _montar_panel_gerencial(
    os_obj,
    producao,
    unidade,
    dados_imovel,
    entrada_dai,
    entrada_eav,
    *,
    request=None,
    servidor_logado=None,
    perfil_pode_homologar=False,
    modo_b=False,
    etapa_interna=None,
    etapa_interna_choices=None,
    os_editavel=False,
    pode_criar_producao=False,
    processos_list=None,
    macroetapa=None,
    status_unidade=None,
):
    from core.os_service import ETAPAS_INTERNAS_LABELS

    dias_sei = _calcular_dias_sei_gerencial(os_obj)
    macroetapa = macroetapa or macroetapa_atual_os(os_obj)
    processos = processos_list or []
    processo_principal = (
        os_obj.processos_vinculados.filter(tipo_vinculo="PRINCIPAL")
        .select_related("processo_sei")
        .first()
    )
    processo_sei = "—"
    if processo_principal and processo_principal.processo_sei:
        processo_sei = processo_principal.processo_sei.numero_processo

    servidores_unidade = []
    if unidade:
        servidores_unidade = [
            {"pk": s.pk, "nome": s.nome}
            for s in Servidor.objects.filter(
                vinculos_unidade__unidade=unidade,
                vinculos_unidade__data_fim__isnull=True,
            )
            .distinct()
            .order_by("nome")
        ]

    comentarios_os_qs = (
        Comentario.objects.filter(os=os_obj, origem="OS")
        .select_related("servidor")
        .order_by("-data_hora")
    )

    return {
        "os_pk": os_obj.pk,
        "producao_pk": producao.pk if producao else None,
        "producao_pk_ativa": producao.pk if producao else None,
        "numero_os": os_obj.numero_os,
        "processo_sei": processo_sei,
        "processos": processos,
        "macroetapa": macroetapa,
        "macroetapa_label": _label_macroetapa_gerencial(macroetapa),
        "macroetapa_cor": _cor_macroetapa_gerencial(macroetapa),
        "etapa_interna": etapa_interna or "",
        "etapa_interna_label": ETAPAS_INTERNAS_LABELS.get(
            etapa_interna or "",
            etapa_interna or "—",
        ),
        "etapa_interna_choices": etapa_interna_choices or [],
        "status_unidade": status_unidade,
        "os_editavel": os_editavel,
        "pode_criar_producao": pode_criar_producao,
        "producoes": _producoes_painel_gerencial(
            os_obj,
            unidade,
            servidor_logado,
            perfil_pode_homologar,
            request,
            producao_ativa=producao,
            modo_b=modo_b,
        ),
        "servidores_unidade": servidores_unidade,
        "comentarios_os": [
            _serializar_comentario(c) for c in comentarios_os_qs[:5]
        ],
        "total_comentarios_os": comentarios_os_qs.count(),
        "entrada_eav": entrada_eav,
        "dias_sei": dias_sei,
        "apelido": os_obj.apelido or "",
        "modo_b": modo_b,
    }


ETAPAS_FLUXO_GERENCIAL = ["ENTRADA", "TRIAGEM", "EM_ATENDIMENTO"]


def _etapas_posteriores_gerencial(etapa_atual):
    from core.os_service import ETAPAS_INTERNAS_LABELS

    if not etapa_atual:
        posteriores = list(ETAPAS_FLUXO_GERENCIAL)
    elif etapa_atual == "EM_ATENDIMENTO":
        return [
            {
                "valor": "CONCLUIDA",
                "label": ETAPAS_INTERNAS_LABELS.get("CONCLUIDA", "Concluída"),
            },
        ]
    else:
        try:
            idx = ETAPAS_FLUXO_GERENCIAL.index(etapa_atual)
            posteriores = ETAPAS_FLUXO_GERENCIAL[idx + 1 :]
        except ValueError:
            posteriores = list(ETAPAS_FLUXO_GERENCIAL)
    return [
        {"valor": etapa, "label": ETAPAS_INTERNAS_LABELS.get(etapa, etapa)}
        for etapa in posteriores
    ]


def _pode_criar_producao_gerencial(os_obj, status_unidade):
    if status_unidade not in ("ABERTA", "REABERTA"):
        return False
    return not (
        os_obj.producoes.exclude(
            status__in=[Producao.STATUS_ENVIADO, Producao.STATUS_CANCELADO],
        ).exists()
    )


def _serializar_linhas_gerencial(
    os_obj,
    unidade_logada=None,
    servidor_logado=None,
    perfil_pode_homologar=False,
    request=None,
    modo_b=False,
):
    """Retorna lista de linhas gerenciais (uma por produção, ou sem produção)."""
    from core.os_service import ETAPAS_INTERNAS_LABELS

    processos = os_obj.processos_vinculados.select_related(
        "processo_sei",
    ).order_by("tipo_vinculo", "data_entrada_divisao")

    numeros_processos = [
        {
            "numero": op.processo_sei.numero_processo,
            "tipo_vinculo": op.tipo_vinculo,
            "data_entrada_divisao": (
                op.data_entrada_divisao.strftime("%d/%m/%Y")
                if op.data_entrada_divisao
                else "—"
            ),
        }
        for op in processos
        if op.processo_sei_id
    ]

    macroetapa = macroetapa_atual_os(os_obj)
    status_unidade = None
    etapa_interna = None
    etapa_interna_label = None
    if unidade_logada:
        status_unidade = (
            OsUnidadeStatus.objects.filter(
                os=os_obj,
                unidade=unidade_logada,
            )
            .values_list("status", flat=True)
            .first()
        )
        if status_unidade in ("ABERTA", "REABERTA"):
            tarefa = (
                TarefaInterna.objects.filter(
                    os=os_obj,
                    unidade=unidade_logada,
                )
                .order_by("-data_inicio")
                .first()
            )
            if tarefa:
                etapa_interna = tarefa.etapa_interna
                etapa_interna_label = ETAPAS_INTERNAS_LABELS.get(
                    etapa_interna,
                    etapa_interna,
                )

    os_editavel = False
    if request is not None:
        os_editavel = os_editavel_para_usuario(os_obj, request)

    total_comentarios = Comentario.objects.filter(os=os_obj).count()
    etapa_interna_choices = _etapas_posteriores_gerencial(etapa_interna)
    pode_criar_producao = _pode_criar_producao_gerencial(os_obj, status_unidade)

    processo_principal = next(
        (p for p in processos if p.tipo_vinculo == "PRINCIPAL"),
        None,
    )
    if processo_principal and processo_principal.data_entrada_divisao:
        entrada_dai = processo_principal.data_entrada_divisao
    else:
        entrada_dai = os_obj.data_entrada_divisao
    entrada_unidade = (
        data_entrada_unidade(os_obj, unidade_logada) if unidade_logada else None
    )
    entrada_eav = (
        timezone.localtime(entrada_unidade).strftime("%d/%m/%Y %H:%M")
        if entrada_unidade
        else "—"
    )
    dias_sei = _calcular_dias_sei_gerencial(os_obj)
    prioridade_label = PRIORIDADE_OS_LABELS.get(
        os_obj.prioridade,
        os_obj.prioridade or "—",
    )

    dados_fixos = {
        "os_pk": os_obj.pk,
        "numero_os": os_obj.numero_os,
        "apelido": os_obj.apelido or "",
        "pendente_encaminhamento": bool(os_obj.pendente_encaminhamento),
        "processos": numeros_processos,
        "macroetapa": macroetapa,
        "macroetapa_label": _label_macroetapa_gerencial(macroetapa),
        "etapa_interna": etapa_interna,
        "etapa_interna_label": etapa_interna_label or etapa_interna or "",
        "etapa_interna_choices": etapa_interna_choices,
        "entrada_dai": _formatar_data_br(entrada_dai),
        "entrada_eav": entrada_eav,
        "requerimento": os_obj.tipo_demanda.descricao,
        "finalidade": os_obj.finalidade.descricao,
        "prioridade": prioridade_label,
        "dias_sei": dias_sei,
        "status_unidade": status_unidade,
        "os_editavel": os_editavel,
        "total_comentarios": total_comentarios,
        "pode_criar_producao": pode_criar_producao,
        "etapa_interna_choices_json": json.dumps(
            etapa_interna_choices,
            ensure_ascii=False,
        ),
    }

    outra_equipe = False
    if unidade_logada:
        servidores_unidade = ServidorUnidade.objects.filter(
            unidade=unidade_logada,
            data_fim__isnull=True,
        ).values_list("servidor_id", flat=True)

        producoes = (
            os_obj.producoes.exclude(status=Producao.STATUS_CANCELADO)
            .filter(
                Q(unidade=unidade_logada)
                | Q(criado_por_id__in=servidores_unidade),
            )
            .select_related("tipo_producao")
            .order_by("-data_criacao")
        )

        producoes = list(producoes)
        if not producoes and os_obj.producoes.exclude(
            status=Producao.STATUS_CANCELADO,
        ).exists():
            outra_equipe = True
    else:
        producoes = list(
            os_obj.producoes.exclude(status=Producao.STATUS_CANCELADO)
            .select_related("tipo_producao")
            .order_by("-data_criacao")
        )

    linhas = []

    def finalizar_linha(linha, producao=None, dados_imovel=None):
        imovel = dados_imovel or _campos_imovel_vazios_gerencial()
        linha["cells"] = _montar_cells_gerencial(linha)
        linha["panel"] = _montar_panel_gerencial(
            os_obj,
            producao,
            unidade_logada,
            imovel,
            entrada_dai,
            entrada_eav,
            request=request,
            servidor_logado=servidor_logado,
            perfil_pode_homologar=perfil_pode_homologar,
            modo_b=modo_b,
            etapa_interna=etapa_interna,
            etapa_interna_choices=etapa_interna_choices,
            os_editavel=os_editavel,
            pode_criar_producao=pode_criar_producao,
            processos_list=numeros_processos,
            macroetapa=macroetapa,
            status_unidade=status_unidade,
        )
        linha["panel_json"] = json.dumps(linha["panel"], default=str)
        return linha

    if outra_equipe:
        dados_imovel = _dados_imovel_gerencial(os_obj)
        linha = {**dados_fixos, **_campos_vazios_gerencial(), **dados_imovel}
        linha.update(
            {
                "tem_producao": False,
                "status_producao": "",
                "status_producao_label": "Outra equipe",
                "status_producao_cor": "secondary",
                "apelido": os_obj.apelido or "",
                "prioridade": prioridade_label,
                "dias_sei": dias_sei,
            },
        )
        linhas.append(finalizar_linha(linha, None, dados_imovel))
        return linhas

    if not producoes:
        dados_imovel = _dados_imovel_gerencial(os_obj)
        linha = {**dados_fixos, **_campos_vazios_gerencial(), **dados_imovel}
        linha.update(
            {
                "tem_producao": False,
                "status_producao": "",
                "status_producao_label": "Sem produção",
                "status_producao_cor": "secondary",
                "apelido": os_obj.apelido or "",
                "prioridade": prioridade_label,
                "dias_sei": dias_sei,
            },
        )
        linhas.append(finalizar_linha(linha, None, dados_imovel))
        return linhas

    for producao in producoes:
        dados_imovel = _dados_imoveis_producao_gerencial(producao)
        linha = {**dados_fixos}
        linha.update(
            {
                "tem_producao": True,
                "producao_pk": producao.pk,
                "status_producao": producao.status,
                "status_producao_label": _label_status_producao_gerencial(
                    producao.status,
                ),
                "status_producao_cor": _cor_status_producao_gerencial(producao.status),
                **dados_imovel,
                "apelido": os_obj.apelido or "",
                "prioridade": prioridade_label,
                "dias_sei": dias_sei,
                "enviado": _formatar_data_br(producao.data_enviado),
                "la_pt_ptf": producao.numero_producao or "—",
                "tipo_trabalho": (
                    producao.tipo_producao.label_display
                    if producao.tipo_producao
                    else "—"
                ),
                "doc_sei": producao.numero_sei or "—",
                "destino": _destino_pos_homologacao(os_obj, producao),
            },
        )
        linhas.append(finalizar_linha(linha, producao, dados_imovel))

    return linhas


def _montar_linhas_gerencial(
    os_queryset,
    unidade,
    servidor_logado=None,
    perfil_pode_homologar=False,
    request=None,
    modo_b=False,
):
    os_list = list(
        os_queryset.select_related("natureza", "tipo_demanda", "finalidade"),
    )
    if not os_list:
        return []

    linhas = []
    for os_obj in os_list:
        linhas.extend(
            _serializar_linhas_gerencial(
                os_obj,
                unidade,
                servidor_logado=servidor_logado,
                perfil_pode_homologar=perfil_pode_homologar,
                request=request,
                modo_b=modo_b,
            ),
        )
    return linhas


def _filtrar_linhas_coluna_gerencial(linhas, request):
    filtradas = linhas
    for coluna in COLUNAS_GERENCIAL_CONFIG:
        valor = (request.GET.get(f"fg_{coluna}") or "").strip()
        if not valor:
            continue
        if coluna == "dias_sei":
            filtradas = [
                linha
                for linha in filtradas
                if linha["cells"].get("dias_sei") is not None
                and str(linha["cells"]["dias_sei"]) == valor
            ]
        else:
            filtradas = [
                linha
                for linha in filtradas
                if linha["cells"].get(coluna) == valor
            ]
    return filtradas


def _opcoes_filtro_coluna_gerencial(linhas):
    opcoes = {}
    for coluna in COLUNAS_GERENCIAL_CONFIG:
        valores = set()
        for linha in linhas:
            valor = linha["cells"].get(coluna)
            if coluna == "dias_sei":
                if valor is not None:
                    valores.add(str(valor))
            elif valor and valor != "—":
                valores.add(valor)
        opcoes[coluna] = sorted(valores, key=lambda item: (item == "—", item))
    return opcoes


def _contagens_status_gerencial(os_ids):
    contagens = {status: 0 for status in STATUS_GERENCIAL_CARDS}
    if not os_ids:
        return contagens
    for item in (
        Producao.objects.filter(os_id__in=os_ids)
        .values("status")
        .annotate(total=Count("id"))
    ):
        if item["status"] in contagens:
            contagens[item["status"]] = item["total"]
    return contagens


def _contexto_gerencial_os_list(request, queryset_completo, linhas_pagina, modo_b=False):
    servidor = _obter_servidor(request.user)
    unidade = _obter_unidade_principal_servidor(servidor)
    perfil = getattr(request, "perfil_acesso", None)
    pode_homologar = perfil is not None and perfil.pode_homologar
    os_ids = list(queryset_completo.values_list("pk", flat=True))
    linhas_base = _montar_linhas_gerencial(
        queryset_completo,
        unidade,
        servidor_logado=servidor,
        perfil_pode_homologar=pode_homologar,
        request=request,
        modo_b=modo_b,
    )
    linhas_filtradas = _filtrar_linhas_coluna_gerencial(linhas_base, request)
    os_ids_filtrados = {linha["os_pk"] for linha in linhas_filtradas}

    filtros_coluna_ativos = {
        coluna: request.GET.get(f"fg_{coluna}", "").strip()
        for coluna in COLUNAS_GERENCIAL_CONFIG
        if request.GET.get(f"fg_{coluna}", "").strip()
    }

    colunas_visiveis = _colunas_visiveis_gerencial(servidor)

    return {
        "linhas_gerencial": linhas_pagina,
        "colunas_gerencial_config": COLUNAS_GERENCIAL_CONFIG,
        "colunas_gerencial_labels": {
            chave: meta["label"] for chave, meta in COLUNAS_GERENCIAL_CONFIG.items()
        },
        "colunas_gerencial_visiveis": colunas_visiveis,
        "colunas_gerencial_visiveis_json": json.dumps(colunas_visiveis),
        "grupos_colunas_gerencial": _grupos_header_gerencial(colunas_visiveis),
        "contagens_status_gerencial": _contagens_status_gerencial(os_ids),
        "opcoes_filtro_coluna": _opcoes_filtro_coluna_gerencial(linhas_base),
        "filtros_coluna_ativos": filtros_coluna_ativos,
        "status_gerencial_cards": STATUS_GERENCIAL_CARDS,
        "query_string_base": _query_string_os_list(request),
        "query_string_gerencial_base": _query_string_gerencial(request),
        "urls_status_gerencial": {
            status: _query_string_gerencial(request, status_producao=status)
            for status in STATUS_GERENCIAL_CARDS
        },
        "gerencial_panels_json": json.dumps(
            {
                (
                    str(linha["producao_pk"])
                    if linha.get("producao_pk")
                    else f"os-{linha['os_pk']}"
                ): linha["panel"]
                for linha in linhas_pagina
            },
            default=str,
            ensure_ascii=False,
        ),
        "pode_homologar": pode_homologar,
        "pode_editar_entrada_dai": _pode_editar_entrada_dai(request),
        "servidores": Servidor.objects.order_by("nome"),
        "servidores_unidade": (
            Servidor.objects.filter(
                vinculos_unidade__unidade=unidade,
                vinculos_unidade__data_fim__isnull=True,
            )
            .distinct()
            .order_by("nome")
            if unidade
            else Servidor.objects.none()
        ),
        "tipos_producao_unidade": (
            TipoProducao.objects.filter(
                ativo=True,
                unidades_competentes__unidade_interna=unidade,
            )
            .distinct()
            .order_by("prefixo", "subtipo")
            if unidade
            else TipoProducao.objects.none()
        ),
        "prioridades_os": ["NORMAL", "PRIORITARIO", "URGENTE"],
        "prazo_tipo_opcoes": OS.PRAZO_TIPO_CHOICES,
        "status_producao_opcoes_gerencial": Producao.STATUS_CHOICES,
        "os_ids_filtrados_count": len(os_ids_filtrados),
        "gerencial_modo_b": modo_b,
    }


def _etapas_unidades_abertas(os):
    from core.os_service import ETAPAS_INTERNAS_LABELS

    status_abertos = OsUnidadeStatus.objects.filter(
        os=os,
        status__in=("ABERTA", "REABERTA"),
    ).select_related("unidade")

    resultado = []
    for su in status_abertos:
        tarefa = (
            TarefaInterna.objects.filter(
                os=os,
                unidade=su.unidade,
                status="PENDENTE",
            )
            .order_by("-data_inicio")
            .first()
        )

        if tarefa and tarefa.etapa_interna:
            etapa = ETAPAS_INTERNAS_LABELS.get(
                tarefa.etapa_interna,
                tarefa.etapa_interna,
            )
        else:
            etapa = "—"

        resultado.append(
            {
                "sigla": su.unidade.sigla,
                "etapa": etapa,
                "status_unidade": su.status,
            }
        )
    return resultado


def _enriquecer_os_lista(ordens):
    from core.templatetags.siprac_filters import MACROETAPA_LABELS

    for os_obj in ordens:
        os_obj.unidades_abertas = _etapas_unidades_abertas(os_obj)
        os_obj.macroetapa_display = MACROETAPA_LABELS.get(
            getattr(os_obj, "macroetapa_atual", None),
            getattr(os_obj, "macroetapa_atual", None) or "—",
        )
        processos = list(os_obj.processos_vinculados.all())
        processos.sort(key=lambda op: (0 if op.tipo_vinculo == "PRINCIPAL" else 1, op.pk))
        os_obj.processos_lista = [
            {
                "numero_processo": op.processo_sei.numero_processo,
                "tipo_vinculo": op.tipo_vinculo,
            }
            for op in processos
        ]


class OSListView(RequerLoginMixin, ListView):
    template_name = "os_list.html"
    context_object_name = "ordens"
    paginate_by = 20

    def get_paginate_by(self, queryset):
        if self.request.GET.get("view") in ("gerencial", "gerencial_b"):
            return 50
        return self.paginate_by

    def get_queryset(self):
        os_visiveis = _queryset_os_por_visibilidade(self.request)
        queryset = _queryset_os_anotado().filter(pk__in=os_visiveis.values("pk"))
        queryset = _aplicar_filtros_os(queryset, self.request)
        self._qs_gerencial_completo = queryset

        if self.request.GET.get("view") in ("gerencial", "gerencial_b"):
            servidor = _obter_servidor(self.request.user)
            unidade = _obter_unidade_principal_servidor(servidor)
            perfil = getattr(self.request, "perfil_acesso", None)
            pode_homologar = perfil is not None and perfil.pode_homologar
            linhas = _montar_linhas_gerencial(
                queryset,
                unidade,
                servidor_logado=servidor,
                perfil_pode_homologar=pode_homologar,
                request=self.request,
                modo_b=(self.request.GET.get("view") == "gerencial_b"),
            )
            linhas = _filtrar_linhas_coluna_gerencial(linhas, self.request)
            os_ids = [linha["os_pk"] for linha in linhas]
            if os_ids:
                queryset = queryset.filter(pk__in=os_ids)
            else:
                queryset = queryset.none()
        else:
            queryset = queryset.prefetch_related(
                "processos_vinculados__processo_sei",
            )

        return _ordenar_queryset_os_fila(queryset)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        view_mode = self.request.GET.get("view", "lista")
        context["view_mode"] = view_mode
        context["filtro_macroetapa"] = self.request.GET.get("macroetapa", "")
        context["filtro_natureza"] = self.request.GET.get("natureza", "")
        context["filtro_prioridade"] = self.request.GET.get("prioridade", "")
        context["filtro_q"] = self.request.GET.get("q", "")
        context["filtro_status_producao"] = self.request.GET.get("status_producao", "")
        context["filtro_aguarda_retorno"] = self.request.GET.get("aguarda_retorno", "")
        context["filtro_prazo"] = self.request.GET.get("prazo", "")
        context["filtro_unidade"] = self.request.GET.get("unidade", "")
        context["naturezas"] = Natureza.objects.filter(ativa=True).order_by("descricao")
        context["status_producao_opcoes"] = Producao.STATUS_CHOICES
        context["macroetapas"] = [
            choice[0]
            for choice in Encaminhamento.TIPO_MACROETAPA_CHOICES
        ] + ["ENCERRADO"]
        context["prioridades"] = ["NORMAL", "PRIORITARIO", "URGENTE"]
        context["query_string_lista"] = _query_string_os_list(
            self.request,
            view="lista",
        )
        context["query_string_gerencial"] = _query_string_os_list(
            self.request,
            view="gerencial",
        )
        context["query_string_gerencial_b"] = _query_string_os_list(
            self.request,
            view="gerencial_b",
        )
        context["query_string_filtros"] = _query_string_filtros_os_list(self.request)

        if view_mode in ("gerencial", "gerencial_b"):
            servidor = _obter_servidor(self.request.user)
            unidade = _obter_unidade_principal_servidor(servidor)
            perfil = getattr(self.request, "perfil_acesso", None)
            pode_homologar = perfil is not None and perfil.pode_homologar
            linhas_pagina = _montar_linhas_gerencial(
                context["ordens"],
                unidade,
                servidor_logado=servidor,
                perfil_pode_homologar=pode_homologar,
                request=self.request,
                modo_b=(view_mode == "gerencial_b"),
            )
            context.update(
                _contexto_gerencial_os_list(
                    self.request,
                    getattr(self, "_qs_gerencial_completo", self.get_queryset()),
                    linhas_pagina,
                    modo_b=(view_mode == "gerencial_b"),
                ),
            )
        else:
            _enriquecer_os_lista(context["ordens"])
        return context


class ProducaoListView(RequerLoginMixin, ListView):
    template_name = "producao_list.html"
    context_object_name = "producoes"
    paginate_by = 20

    def get_queryset(self):
        queryset = Producao.objects.select_related(
            "os",
            "tipo_producao",
        )
        queryset = _aplicar_visibilidade_producao(queryset, self.request)
        queryset = _aplicar_filtros_producao(queryset, self.request)
        return queryset.order_by("-data_criacao")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filtro_status"] = self.request.GET.get("status", "")
        context["filtro_tipo_producao"] = self.request.GET.get("tipo_producao", "")
        context["filtro_periodo"] = self.request.GET.get("periodo", "")
        context["filtro_unidade"] = self.request.GET.get("unidade", "")
        context["status_opcoes"] = Producao.STATUS_CHOICES
        context["tipos_producao"] = TipoProducao.objects.filter(ativo=True).order_by(
            "prefixo",
        )
        context["unidades"] = UnidadeInterna.objects.all().order_by("sigla")
        perfil = getattr(self.request, "perfil_acesso", None)
        context["exibir_filtro_unidade"] = perfil and perfil.visibilidade_total
        return context


class ProcessoDetailView(RequerLoginMixin, DetailView):
    model = ProcessoSei
    template_name = "processo_detail.html"
    context_object_name = "processo"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        processo = self.object

        vinculos = (
            OsProcesso.objects.filter(processo_sei=processo)
            .select_related("os")
            .order_by("-data_entrada_divisao", "os__numero_os")
        )
        os_ids = [vinculo.os_id for vinculo in vinculos]
        macroetapas = {}
        if os_ids:
            for os_obj in _queryset_os_anotado().filter(pk__in=os_ids):
                macroetapas[os_obj.pk] = os_obj.macroetapa_atual

        os_vinculadas = []
        for vinculo in vinculos:
            os_vinculadas.append(
                {
                    "os_pk": vinculo.os_id,
                    "numero_os": vinculo.os.numero_os,
                    "tipo_vinculo": vinculo.tipo_vinculo,
                    "data_entrada": vinculo.data_entrada_divisao,
                    "data_encerramento": vinculo.data_encerramento,
                    "macroetapa": macroetapas.get(vinculo.os_id),
                },
            )

        producoes = (
            Producao.objects.filter(os_id__in=os_ids)
            .select_related("tipo_producao")
            .order_by("-data_criacao")
        ) if os_ids else Producao.objects.none()

        context["os_vinculadas"] = os_vinculadas
        context["producoes"] = producoes
        context["outras_os_mesmo_imovel"] = _obter_outras_os_mesmo_imovel(processo)
        return context


class OSDetailView(RequerLoginMixin, DetailView):
    model = OS
    template_name = "os_detail.html"
    context_object_name = "os"

    def get_queryset(self):
        return super().get_queryset().select_related(
            "natureza",
            "tipo_demanda",
            "finalidade",
            "criado_por",
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        os_obj = self.object

        context["processos"] = (
            OsProcesso.objects.filter(os=os_obj)
            .select_related("processo_sei")
            .order_by("-data_vinculo")
        )
        context["processo_principal"] = (
            OsProcesso.objects.filter(os=os_obj, tipo_vinculo="PRINCIPAL")
            .select_related("processo_sei")
            .first()
        )
        os_imoveis = OsImovel.objects.filter(os=os_obj)
        imoveis_em_producao = set(
            ProducaoImovel.objects.filter(
                os_imovel__os=os_obj,
            )
            .exclude(
                producao__status=Producao.STATUS_CANCELADO,
            )
            .values_list("os_imovel_id", flat=True),
        )
        context["imoveis_sem_producao"] = {
            oi.pk for oi in os_imoveis if oi.pk not in imoveis_em_producao
        }
        context["timeline"] = timeline_os(os_obj)
        imoveis_vinculados = OsImovel.objects.filter(os=os_obj).select_related(
            "imovel",
            "vinculado_por",
        )
        context["imoveis"] = imoveis_vinculados
        imoveis_com_coords, imoveis_sem_coordenadas = _montar_imoveis_coords_os(
            imoveis_vinculados,
        )
        context["imoveis_com_coords"] = imoveis_com_coords
        context["imoveis_sem_coordenadas"] = imoveis_sem_coordenadas
        context["imoveis_coords_json"] = json.dumps(
            imoveis_com_coords,
            ensure_ascii=False,
        )
        context["producoes"] = (
            Producao.objects.filter(os=os_obj)
            .select_related("tipo_producao")
            .order_by("-data_criacao")
        )
        context["macroetapa_atual"] = macroetapa_atual_os(os_obj)
        context["os_encerrada"] = _os_esta_encerrada(os_obj)
        context["tem_servidor"] = _obter_servidor(self.request.user) is not None
        context["comentarios"] = (
            Comentario.objects.filter(os=os_obj)
            .select_related("servidor", "producao", "producao__tipo_producao")
            .order_by("-data_hora")
        )
        servidor = _obter_servidor(self.request.user)
        vinculo = getattr(self.request, "vinculo_ativo", None)
        unidade = (
            vinculo.unidade
            if vinculo
            else (_obter_unidade_principal_servidor(servidor) if servidor else None)
        )
        context["data_entrada_unidade_atual"] = (
            data_entrada_unidade(os_obj, unidade) if unidade else None
        )
        status_unidade = None
        if vinculo:
            status_unidade = OsUnidadeStatus.objects.filter(
                os=os_obj,
                unidade=vinculo.unidade,
            ).first()
        elif unidade:
            status_unidade = OsUnidadeStatus.objects.filter(
                os=os_obj,
                unidade=unidade,
            ).first()
        context["status_unidade"] = status_unidade
        context["status_unidade_atual"] = status_unidade
        context["os_editavel"] = os_editavel_para_usuario(os_obj, self.request)
        context["is_primeiro_encaminhamento"] = is_primeiro_encaminhamento(os_obj)
        perfil = getattr(self.request, "perfil_acesso", None)
        context["pode_homologar"] = bool(perfil and perfil.pode_homologar)
        context["pode_reabrir_na_unidade"] = (
            perfil is not None
            and perfil.pode_homologar
            and status_unidade is not None
            and status_unidade.status == "CONCLUIDA"
        )
        return context


class EncaminhamentoCreateView(RequerLoginMixin, FormView):
    template_name = "encaminhamento_form.html"
    form_class = EncaminhamentoForm

    def dispatch(self, request, *args, **kwargs):
        self.os_obj = get_object_or_404(OS, pk=kwargs["pk"])
        if _obter_servidor(request.user) is None:
            messages.error(request, MSG_SEM_PERMISSAO)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))
        if not os_editavel_para_usuario(self.os_obj, request):
            messages.error(request, MSG_OS_SOMENTE_LEITURA)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))
        if getattr(request, "visibilidade", "UNIDADE") == "DEPARTAMENTO":
            # DEPARTAMENTO só pode fazer o primeiro encaminhamento
            # ou incluir processo relacionado
            if not is_primeiro_encaminhamento(self.os_obj):
                messages.error(
                    request,
                    "Perfil DEPARTAMENTO só pode fazer o primeiro "
                    "encaminhamento. Para reencaminhar, use um "
                    "vínculo operacional.",
                )
                return redirect(
                    reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}),
                )
        return super().dispatch(request, *args, **kwargs)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["unidade_interna_destino"].queryset = (
            UnidadeInterna.objects.filter(tipo="OPERACIONAL").order_by("sigla")
        )
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["os"] = self.os_obj
        servidor = _obter_servidor(self.request.user)
        context["unidade_origem"] = origem_encaminhamento(
            self.os_obj,
            servidor_logado=servidor,
        )
        unidade = _obter_unidade_principal_servidor(servidor) if servidor else None
        context["mostrar_manter_aberta"] = bool(
            unidade
            and OsUnidadeStatus.objects.filter(
                os=self.os_obj,
                unidade=unidade,
                status="ABERTA",
            ).exists()
        )
        unidade_atual = unidade_atual_da_os(self.os_obj)
        context["unidade_atual_pk"] = unidade_atual.pk if unidade_atual else None
        return context

    def get_initial(self):
        initial = super().get_initial()
        # Deixa etapa em branco; o JS preenche conforme o destino selecionado.
        initial.setdefault("etapa_interna", "")
        return initial

    def form_valid(self, form):
        servidor = _obter_servidor(self.request.user)
        if servidor is None:
            messages.error(self.request, MSG_SEM_PERMISSAO)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))

        vinculo = obter_vinculo_unidade_ativo(servidor)
        if vinculo is None:
            form.add_error(None, "Servidor sem vínculo ativo em unidade.")
            return self.form_invalid(form)

        agora = timezone.now()
        dados = form.cleaned_data
        tipo_destino = dados["tipo_destino"]
        unidade_origem = origem_encaminhamento(
            self.os_obj,
            servidor_logado=servidor,
        )
        manter_aberta = bool(dados.get("manter_aberta_na_unidade"))
        if tipo_destino == "EXTERNO":
            etapa_interna = None
            etapa_tarefa = "TRIAGEM"
        else:
            etapa_interna = dados["etapa_interna"]
            etapa_tarefa = etapa_interna

        with transaction.atomic():
            encaminhamento = Encaminhamento.objects.create(
                os=self.os_obj,
                unidade_interna_origem=unidade_origem,
                servidor_origem=servidor,
                unidade_interna_destino=dados.get("unidade_interna_destino"),
                servidor_destino=dados.get("servidor_destino"),
                unidade_externa_destino=dados.get("unidade_externa_destino"),
                etapa_interna=etapa_interna,
                tipo_macroetapa=(
                    Encaminhamento.TIPO_MACROETAPA_ATENDIMENTO_INTERNO
                    if tipo_destino == "INTERNO"
                    else (
                        Encaminhamento.TIPO_MACROETAPA_ATENDIMENTO_EXTERNO
                        if dados.get("aguarda_retorno")
                        else None
                    )
                ),
                aguarda_retorno=dados.get("aguarda_retorno") or False,
                data_retorno_prevista=dados.get("data_retorno_prevista"),
                observacao=dados.get("observacao") or None,
                manter_aberta_na_unidade=manter_aberta,
            )

            if tipo_destino == "INTERNO":
                unidade_tarefa = dados["unidade_interna_destino"]
                servidor_tarefa = dados.get("servidor_destino") or servidor
            else:
                unidade_tarefa = vinculo.unidade
                servidor_tarefa = servidor

            TarefaInterna.objects.create(
                os=self.os_obj,
                encaminhamento=encaminhamento,
                unidade=unidade_tarefa,
                servidor=servidor_tarefa,
                etapa_interna=etapa_tarefa,
                status="PENDENTE",
                data_inicio=agora,
            )

            _atualizar_status_unidade_encaminhamento(
                self.os_obj,
                unidade_origem,
                servidor,
                manter_aberta,
                unidade_destino=dados.get("unidade_interna_destino"),
            )
            if self.os_obj.pendente_encaminhamento:
                self.os_obj.pendente_encaminhamento = False
                self.os_obj.save(update_fields=["pendente_encaminhamento"])

        messages.success(self.request, "Encaminhamento registrado.")
        return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))


class OSReabrirNaUnidadeView(RequerLoginMixin, View):
    def post(self, request, pk):
        os_obj = get_object_or_404(OS, pk=pk)
        perfil = getattr(request, "perfil_acesso", None)
        if perfil is None or not perfil.pode_homologar:
            messages.error(request, MSG_SEM_PERMISSAO)
            return redirect(reverse("os_detalhe", kwargs={"pk": os_obj.pk}))

        servidor = _obter_servidor(request.user)
        unidade = _obter_unidade_principal_servidor(servidor) if servidor else None
        if servidor is None or unidade is None:
            messages.error(request, MSG_SEM_PERMISSAO)
            return redirect(reverse("os_detalhe", kwargs={"pk": os_obj.pk}))

        status_unidade = OsUnidadeStatus.objects.filter(
            os=os_obj,
            unidade=unidade,
            status="CONCLUIDA",
        ).first()
        if status_unidade is None:
            messages.error(
                request,
                "A OS não está concluída nesta unidade para reabertura.",
            )
            return redirect(reverse("os_detalhe", kwargs={"pk": os_obj.pk}))

        status_unidade.status = "REABERTA"
        status_unidade.data_conclusao = None
        status_unidade.aberta_por = servidor
        status_unidade.save(
            update_fields=["status", "data_conclusao", "aberta_por"],
        )
        messages.success(request, "OS reaberta nesta unidade.")
        return redirect(reverse("os_detalhe", kwargs={"pk": os_obj.pk}))


class ProducaoCreateView(RequerLoginMixin, FormView):
    template_name = "producao_form.html"
    form_class = ProducaoForm

    def dispatch(self, request, *args, **kwargs):
        self.os_obj = get_object_or_404(OS, pk=kwargs["pk"])
        if _obter_servidor(request.user) is None:
            messages.error(request, MSG_SEM_PERMISSAO)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))
        if not os_editavel_para_usuario(self.os_obj, request):
            messages.error(request, MSG_OS_SOMENTE_LEITURA)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))
        if getattr(request, "visibilidade", "UNIDADE") == "DEPARTAMENTO":
            messages.error(
                request,
                "Perfil DEPARTAMENTO não pode registrar produção. "
                "Use um vínculo operacional.",
            )
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))
        return super().dispatch(request, *args, **kwargs)

    def _unidade_logada(self):
        vinculo = getattr(self.request, "vinculo_ativo", None)
        if vinculo:
            return vinculo.unidade
        servidor = _obter_servidor(self.request.user)
        if servidor is None:
            return None
        vinculo = obter_vinculo_unidade_ativo(servidor)
        return vinculo.unidade if vinculo else None

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["unidade"] = self._unidade_logada()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["os"] = self.os_obj
        return context

    def form_valid(self, form):
        servidor = _obter_servidor(self.request.user)
        if servidor is None:
            messages.error(self.request, MSG_SEM_PERMISSAO)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))

        dados = form.cleaned_data
        ano = timezone.localdate().year
        tipo_producao = dados["tipo_producao_obj"]
        numero_sei = dados.get("numero_sei") or None
        if dados.get("is_despacho"):
            numero_sei = dados["numero_sei"]

        unidade_logada = self._unidade_logada()

        numero_producao = None
        if not dados.get("is_despacho"):
            numero_producao = _gerar_numero_producao(tipo_producao)

        producao = Producao.objects.create(
            os=self.os_obj,
            tipo_producao=tipo_producao,
            numero_producao=numero_producao,
            numero_sei=numero_sei,
            ano=ano,
            status=Producao.STATUS_ENVIADO,
            data_enviado=timezone.localdate(),
            unidade=unidade_logada,
            criado_por=servidor,
            observacao=dados.get("observacao") or None,
        )

        _criar_producao_status_log(
            producao,
            None,
            Producao.STATUS_ENVIADO,
            servidor,
        )

        ativar_atendimento_interno_se_necessario(producao.os, servidor=servidor)
        # Atualizar etapa interna da unidade para EM_ATENDIMENTO
        if unidade_logada:
            registrar_em_atendimento_na_unidade(
                producao.os,
                unidade_logada,
                servidor=servidor,
            )
        messages.success(self.request, "Produção registrada.")
        return redirect(reverse("producao_detail", kwargs={"pk": producao.pk}))


def _os_esta_encerrada(os_obj):
    return os_obj.encerrada


def _verificar_encerramento_automatico_os(os_obj, servidor):
    processos_abertos = OsProcesso.objects.filter(
        os=os_obj,
        data_encerramento__isnull=True,
    )
    if processos_abertos.exists():
        return False

    if not os_obj.encerrada:
        os_obj.encerrada = True
        os_obj.data_encerramento = timezone.now()
        os_obj.save(update_fields=["encerrada", "data_encerramento"])
        return True
    return False


class OSEncerramentoView(RequerLoginMixin, FormView):
    template_name = "os_encerramento.html"
    form_class = OSEncerramentoForm

    def dispatch(self, request, *args, **kwargs):
        self.os_obj = get_object_or_404(OS, pk=kwargs["pk"])
        perfil = getattr(request, "perfil_acesso", None)
        if perfil is None or not perfil.pode_encerrar_os:
            messages.error(request, MSG_SEM_PERMISSAO)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))
        if not os_editavel_para_usuario(self.os_obj, request):
            messages.error(request, MSG_OS_SOMENTE_LEITURA)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["os"] = self.os_obj
        context["processos_a_encerrar"] = OsProcesso.objects.filter(
            os=self.os_obj,
            data_encerramento__isnull=True,
        ).select_related("processo_sei")
        return context

    def form_valid(self, form):
        servidor = _obter_servidor(self.request.user)
        if servidor is None:
            messages.error(self.request, MSG_SEM_PERMISSAO)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))

        hoje = timezone.localdate()
        motivo = form.cleaned_data["motivo_encerramento"]

        with transaction.atomic():
            self.os_obj.encerrada = True
            self.os_obj.data_encerramento = timezone.now()
            self.os_obj.save(update_fields=["encerrada", "data_encerramento"])
            OsProcesso.objects.filter(
                os=self.os_obj,
                data_encerramento__isnull=True,
            ).update(
                data_encerramento=hoje,
                motivo_encerramento=motivo,
                encerrado_por=servidor,
            )
            _verificar_encerramento_automatico_os(self.os_obj, servidor)

        messages.success(self.request, "OS encerrada com sucesso.")
        return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))


class OSVincularProcessoView(RequerLoginMixin, FormView):
    template_name = "os_vincular_processo.html"
    form_class = OSVincularProcessoForm

    def dispatch(self, request, *args, **kwargs):
        self.os_obj = get_object_or_404(OS, pk=kwargs["pk"])
        if _obter_servidor(request.user) is None:
            messages.error(request, MSG_SEM_PERMISSAO)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))
        if _os_esta_encerrada(self.os_obj):
            messages.error(request, "Não é possível incluir processos em OS encerrada.")
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))
        if not os_editavel_para_usuario(self.os_obj, request):
            messages.error(request, MSG_OS_SOMENTE_LEITURA)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["os"] = self.os_obj
        return context

    def form_valid(self, form):
        servidor = _obter_servidor(self.request.user)
        numero = form.cleaned_data["numero_processo"]
        data_criacao_sei = form.cleaned_data["data_criacao_sei"]
        data_entrada_divisao = form.cleaned_data["data_entrada_divisao"]

        processo_sei, created = ProcessoSei.objects.get_or_create(
            numero_processo=numero,
            defaults={"data_abertura_sei": data_criacao_sei},
        )
        if not created:
            if (
                processo_sei.data_abertura_sei
                and processo_sei.data_abertura_sei != data_criacao_sei
            ):
                messages.warning(
                    self.request,
                    "O processo já existe no sistema com data de criação no SEI "
                    f"diferente ({processo_sei.data_abertura_sei:%d/%m/%Y}). "
                    "O vínculo foi criado com os dados cadastrados anteriormente.",
                )
            elif not processo_sei.data_abertura_sei:
                processo_sei.data_abertura_sei = data_criacao_sei
                processo_sei.save(update_fields=["data_abertura_sei"])

        if OsProcesso.objects.filter(os=self.os_obj, processo_sei=processo_sei).exists():
            messages.error(
                self.request,
                "Este processo já está vinculado a esta OS.",
            )
            return self.form_invalid(form)

        with transaction.atomic():
            os_processo = OsProcesso.objects.create(
                os=self.os_obj,
                processo_sei=processo_sei,
                tipo_vinculo="RELACIONADO",
                data_entrada_divisao=data_entrada_divisao,
            )
            registrar_encaminhamento_automatico(
                self.os_obj,
                Encaminhamento.TIPO_MACROETAPA_INCLUSAO_PROCESSO,
                servidor=servidor,
                observacao=f"Processo {numero} incluído como relacionado.",
            )

            unidade_atual = unidade_atual_da_os(self.os_obj)
            if unidade_atual:
                os_processo.aguardando_redistribuicao = True
                os_processo.observacao_bloqueio = (
                    f"Processo incluído enquanto OS estava em atendimento "
                    f"na unidade {unidade_atual.sigla}. "
                    f"Aguardando redistribuição para esta equipe."
                )
                os_processo.save(
                    update_fields=[
                        "aguardando_redistribuicao",
                        "observacao_bloqueio",
                    ]
                )

        if os_processo.aguardando_redistribuicao:
            messages.warning(
                self.request,
                f"Processo incluído com bloqueio temporário. "
                f"A OS está atualmente na unidade {unidade_atual.sigla}. "
                f"O novo processo ficará bloqueado até ser redistribuído "
                f"para a equipe que está atendendo a OS.",
            )
        else:
            messages.success(
                self.request,
                f"Processo {numero} vinculado à OS com sucesso.",
            )
        return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))


class OSLiberarProcessoView(RequerLoginMixin, View):
    template_name = "os_liberar_processo.html"

    def dispatch(self, request, *args, **kwargs):
        perfil = getattr(request, "perfil_acesso", None)
        if not perfil or not perfil.pode_homologar:
            raise PermissionDenied

        self.os_obj = get_object_or_404(OS, pk=kwargs["os_pk"])
        self.os_processo = get_object_or_404(
            OsProcesso,
            pk=kwargs["proc_pk"],
            os=self.os_obj,
            aguardando_redistribuicao=True,
        )
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, os_pk, proc_pk):
        return render(
            request,
            self.template_name,
            {
                "os": self.os_obj,
                "os_processo": self.os_processo,
            },
        )

    def post(self, request, os_pk, proc_pk):
        servidor = _obter_servidor(request.user)
        if servidor is None:
            messages.error(request, MSG_SEM_PERMISSAO)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))

        vinculo = obter_vinculo_unidade_ativo(servidor)
        if vinculo is None:
            messages.error(request, "Servidor sem vínculo ativo em unidade.")
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))

        agora = timezone.now()
        unidade_atual = unidade_atual_da_os(self.os_obj)
        numero_processo = self.os_processo.processo_sei.numero_processo

        with transaction.atomic():
            self.os_processo.aguardando_redistribuicao = False
            self.os_processo.observacao_bloqueio = None
            self.os_processo.save(
                update_fields=["aguardando_redistribuicao", "observacao_bloqueio"]
            )

            if unidade_atual:
                encaminhamento = Encaminhamento.objects.create(
                    os=self.os_obj,
                    unidade_interna_origem=vinculo.unidade,
                    servidor_origem=servidor,
                    unidade_interna_destino=unidade_atual,
                    etapa_interna="ENTRADA",
                    observacao=(
                        f"Liberação do processo {numero_processo} "
                        f"para atendimento na unidade {unidade_atual.sigla}."
                    ),
                )
                TarefaInterna.objects.create(
                    os=self.os_obj,
                    encaminhamento=encaminhamento,
                    unidade=unidade_atual,
                    servidor=servidor,
                    etapa_interna="ENTRADA",
                    status="PENDENTE",
                    data_inicio=agora,
                )

        messages.success(
            request,
            f"Processo {numero_processo} liberado para atendimento.",
        )
        return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))


class ProximoNumeroAPIView(RequerLoginJSONMixin, View):
    def get(self, request):
        tipo_id = request.GET.get("tipo_id")
        if not tipo_id or tipo_id == DESPACHO_VALUE:
            return JsonResponse({"numero": None})

        try:
            tipo_producao = TipoProducao.objects.get(pk=tipo_id, ativo=True)
        except (TipoProducao.DoesNotExist, ValueError, TypeError):
            return JsonResponse({"error": "Tipo de produção inválido."}, status=404)

        return JsonResponse({"numero": _gerar_numero_producao(tipo_producao)})


class TiposDemandaAPIView(RequerLoginJSONMixin, View):
    def get(self, request):
        natureza_id = request.GET.get("natureza_id")
        if not natureza_id:
            return JsonResponse([], safe=False)

        tipos = list(
            TipoDemanda.objects.filter(
                combinacoes_validas__natureza_id=natureza_id,
                ativa=True,
            )
            .distinct()
            .order_by("descricao")
            .values("id", "descricao")
        )
        return JsonResponse(tipos, safe=False)


class FinalidadesAPIView(RequerLoginJSONMixin, View):
    def get(self, request):
        tipo_demanda_id = request.GET.get("tipo_demanda_id")
        if not tipo_demanda_id:
            return JsonResponse([], safe=False)

        finalidades = list(
            Finalidade.objects.filter(
                combinacoes_validas__tipo_demanda_id=tipo_demanda_id,
                ativa=True,
            )
            .distinct()
            .order_by("descricao")
            .values("id", "descricao")
        )
        return JsonResponse(finalidades, safe=False)


class ImovelMapaView(RequerLoginMixin, TemplateView):
    template_name = "imovel_mapa.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        total_imoveis = Imovel.objects.count()
        imoveis_com_coords = []
        imoveis_sem_coords_lista = []

        for imovel in Imovel.objects.order_by("id"):
            os_imovel = _ultimo_os_imovel(imovel)
            lat, lng = _coords_os_imovel(os_imovel)
            if lat is not None and lng is not None:
                item = _imovel_para_mapa(imovel, os_imovel)
                if item["lat"] is not None and item["lng"] is not None:
                    imoveis_com_coords.append(item)
                    continue
            if len(imoveis_sem_coords_lista) < 50:
                imoveis_sem_coords_lista.append(
                    {
                        "id": imovel.pk,
                        "inscricao_cadastral": imovel.inscricao_cadastral,
                        "codigo_isic": imovel.codigo_isic,
                        "nom_logradouro": os_imovel.nom_logradouro if os_imovel else None,
                        "num_endereco": os_imovel.num_endereco if os_imovel else None,
                        "bairro": os_imovel.bairro if os_imovel else None,
                    },
                )

        total_com_coords = len(imoveis_com_coords)
        total_sem_coords = total_imoveis - total_com_coords
        context["imoveis_com_coords"] = imoveis_com_coords
        context["total_imoveis"] = total_imoveis
        context["total_com_coords"] = total_com_coords
        context["total_sem_coords"] = total_sem_coords
        context["pct_com_coords"] = (
            round(total_com_coords * 100 / total_imoveis, 1)
            if total_imoveis
            else 0
        )
        context["imoveis_coords_json"] = json.dumps(
            imoveis_com_coords,
            ensure_ascii=False,
        )
        context["imoveis_sem_coords_lista"] = imoveis_sem_coords_lista
        return context


class ImovelListView(RequerLoginMixin, ListView):
    model = Imovel
    template_name = "imovel_list.html"
    context_object_name = "imoveis"
    paginate_by = 20

    def get_queryset(self):
        queryset = Imovel.objects.all()
        busca = self.request.GET.get("q", "").strip()
        if busca:
            filtros = Q(codigo_isic__icontains=busca)
            filtros |= Q(os_imoveis__nom_logradouro__icontains=busca)
            filtros |= Q(os_imoveis__bairro__icontains=busca)
            filtros |= Q(os_imoveis__num_endereco__icontains=busca)
            try:
                filtros |= Q(inscricao_cadastral=int(busca))
            except ValueError:
                pass
            queryset = queryset.filter(filtros).distinct()
        return queryset.prefetch_related(
            Prefetch(
                "os_imoveis",
                queryset=OsImovel.objects.order_by("-data_vinculo", "-pk"),
            ),
        ).order_by("-id")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filtro_q"] = self.request.GET.get("q", "")
        return context


class ImovelCreateView(RequerLoginMixin, FormView):
    template_name = "imovel_form.html"
    form_class = ISICForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["proximo_codigo_isic"] = _gerar_codigo_isic()
        return context

    def form_valid(self, form):
        imovel = _salvar_isic_from_form(form.cleaned_data)
        messages.success(self.request, "Imóvel cadastrado com sucesso.")
        return redirect(reverse("imovel_detalhe", kwargs={"pk": imovel.pk}))


class ImovelDetailView(RequerLoginMixin, DetailView):
    model = Imovel
    template_name = "imovel_detail.html"
    context_object_name = "imovel"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        imovel = self.object
        context["vinculos_os"] = (
            OsImovel.objects.filter(imovel=imovel)
            .select_related("os")
            .order_by("-data_vinculo", "-os__data_criacao_sgbd")
        )
        context["vinculos_producao"] = (
            ProducaoImovel.objects.filter(os_imovel__imovel=imovel)
            .select_related(
                "producao",
                "producao__tipo_producao",
                "producao__os",
                "os_imovel",
                "os_imovel__imovel",
            )
            .order_by("-producao__data_criacao")
        )
        return context


class ImovelUpdateView(RequerLoginMixin, FormView):
    template_name = "imovel_edit_form.html"
    form_class = ImovelForm

    def dispatch(self, request, *args, **kwargs):
        self.imovel_obj = get_object_or_404(Imovel, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["imovel"] = self.imovel_obj
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["imovel"] = self.imovel_obj
        context["editando"] = True
        return context

    def form_valid(self, form):
        _salvar_imovel_from_form(form.cleaned_data, self.imovel_obj)
        messages.success(self.request, "Imóvel atualizado com sucesso.")
        return redirect(reverse("imovel_detalhe", kwargs={"pk": self.imovel_obj.pk}))


class ProximoIsicAPIView(RequerLoginJSONMixin, View):
    def get(self, request):
        return JsonResponse({"codigo": _gerar_codigo_isic()})


def _formatar_tamanho_arquivo(tamanho):
    if tamanho >= 1024 * 1024:
        return f"{tamanho / (1024 * 1024):.1f} MB"
    if tamanho >= 1024:
        return f"{tamanho / 1024:.1f} KB"
    return f"{tamanho} bytes"


def _contexto_arquivo_siat():
    status = obter_status_arquivo_siat(SIAT_ARQUIVO_PATH)
    contexto = {
        "arquivo_existe": status.get("disponivel", False),
        "contagem_siat_orfaos": contar_imoveis_siat_orfaos(),
    }
    if not status.get("disponivel"):
        return contexto

    modificado = timezone.localtime(
        datetime.datetime.fromisoformat(status["modificado_em"]),
    )
    contexto.update(
        {
            "arquivo_nome": SIAT_ARQUIVO_PATH.name,
            "arquivo_modificado": modificado,
            "arquivo_tamanho": _formatar_tamanho_arquivo(status["tamanho_bytes"]),
            "arquivo_total_registros": status.get("total_registros", 0),
            "arquivo_data": status.get("data_arquivo"),
        },
    )
    return contexto


class SiatCarregarArquivoView(RequerAdminMixin, FormView):
    template_name = "siat_carregar.html"
    form_class = SiatUploadForm
    success_url = reverse_lazy("siat_carregar")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(_contexto_arquivo_siat())
        context["indice_recarregando"] = self.request.session.pop(
            "siat_indice_recarregando",
            False,
        )
        return context

    def form_valid(self, form):
        SIAT_ARQUIVO_PATH.parent.mkdir(parents=True, exist_ok=True)
        uploaded = form.cleaned_data["arquivo"]
        with open(SIAT_ARQUIVO_PATH, "wb") as destino:
            for chunk in uploaded.chunks():
                destino.write(chunk)

        resultado = carregar_arquivo_siat(SIAT_ARQUIVO_PATH)
        if not resultado.get("valido"):
            if SIAT_ARQUIVO_PATH.exists():
                SIAT_ARQUIVO_PATH.unlink()
            messages.error(
                self.request,
                resultado.get("erro", "Formato de arquivo inválido."),
            )
            return self.form_invalid(form)

        def _recarregar_indice():
            try:
                siat_index.carregar_indice(SIAT_ARQUIVO_PATH)
            except Exception as e:
                logging.getLogger(__name__).error(
                    f"Erro ao recarregar índice SIAT: {e}"
                )

        siat_index.marcar_recarregando()
        threading.Thread(target=_recarregar_indice, daemon=True).start()
        self.request.session["siat_indice_recarregando"] = True
        messages.success(
            self.request,
            "Arquivo SIAT carregado com sucesso. "
            "O índice de busca está sendo atualizado em background "
            "(pode levar 1-2 minutos para buscas por inscrição ficarem disponíveis).",
        )
        return super().form_valid(form)


class SiatStatusView(RequerLoginJSONMixin, View):
    def get(self, request):
        if not getattr(request, "admin_sistema", False):
            return JsonResponse({"error": "Sem permissão."}, status=403)
        return JsonResponse(obter_status_arquivo_siat(SIAT_ARQUIVO_PATH))


class SiatStatusIndexView(RequerLoginJSONMixin, View):
    def get(self, request):
        if not getattr(request, "admin_sistema", False):
            return JsonResponse({"error": "Sem permissão."}, status=403)
        return JsonResponse(siat_index.status_indice())


class SiatLimparImoveisView(RequerAdminMixin, View):
    def dispatch(self, request, *args, **kwargs):
        logger.warning("perfil_acesso: %s", getattr(request, "perfil_acesso", None))
        logger.warning(
            "admin_sistema: %s",
            getattr(getattr(request, "perfil_acesso", None), "admin_sistema", None),
        )
        logger.warning("admin_sistema flag: %s", getattr(request, "admin_sistema", None))
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        return redirect("siat_carregar")

    def post(self, request):
        total = contar_imoveis_siat_orfaos()
        if total == 0:
            messages.info(request, "Nenhum imóvel SIAT órfão para remover.")
            return redirect("siat_carregar")

        removidos = limpar_imoveis_siat_orfaos()
        messages.success(
            request,
            f"{removidos:,} imóvel(is) SIAT não vinculado(s) removido(s).".replace(
                ",",
                ".",
            ),
        )
        return redirect("siat_carregar")


class SiatAtualizarInscricaoView(RequerLoginJSONMixin, View):
    def post(self, request, inscricao):
        imovel = Imovel.objects.filter(
            inscricao_cadastral=inscricao,
            tipo_identificacao="CADASTRAL",
        ).first()
        if imovel is None:
            return JsonResponse(
                {
                    "sucesso": False,
                    "mensagem": "Imóvel cadastral não encontrado.",
                },
                status=404,
            )

        if not SIAT_ARQUIVO_PATH.exists():
            return JsonResponse(
                {
                    "sucesso": False,
                    "mensagem": "Arquivo SIAT não disponível no servidor.",
                },
                status=404,
            )

        if atualizar_inscricao_do_arquivo(imovel, SIAT_ARQUIVO_PATH):
            return JsonResponse(
                {
                    "sucesso": True,
                    "mensagem": (
                        f"Inscrição {inscricao} encontrada na View SIAT. "
                        "Dados cadastrais são persistidos ao vincular o imóvel a uma OS."
                    ),
                },
            )

        return JsonResponse(
            {
                "sucesso": False,
                "mensagem": f"Inscrição {inscricao} não encontrada no arquivo SIAT.",
            },
            status=404,
        )


def _formatar_identificacao_imovel(imovel):
    if imovel.inscricao_cadastral:
        return str(imovel.inscricao_cadastral)
    if imovel.codigo_isic:
        return imovel.codigo_isic
    return f"Imóvel #{imovel.pk}"


def _formatar_endereco_imovel(imovel):
    os_imovel = _ultimo_os_imovel(imovel)
    partes = []
    if os_imovel and os_imovel.nom_logradouro:
        partes.append(os_imovel.nom_logradouro)
    if os_imovel and os_imovel.num_endereco:
        partes.append(os_imovel.num_endereco)
    return ", ".join(partes)


def _formatar_identificacao_vinculo(vinculo):
    return _identificacao_os_imovel(_obter_os_imovel_vinculo(vinculo))


def _formatar_endereco_vinculo(vinculo):
    return _endereco_os_imovel(_obter_os_imovel_vinculo(vinculo))


def _formatar_area_vinculo(vinculo):
    os_imovel = _obter_os_imovel_vinculo(vinculo)
    if os_imovel and os_imovel.area_territorial is not None:
        return str(os_imovel.area_territorial)
    return None


def _serializar_os_imovel(vinculo):
    return {
        "os_imovel_id": vinculo.pk,
        "imovel_id": vinculo.imovel_id,
        "identificacao": _formatar_identificacao_vinculo(vinculo),
        "endereco": _formatar_endereco_vinculo(vinculo),
        "area_territorial": _formatar_area_vinculo(vinculo),
    }


def _serializar_producao_imovel(item):
    return {
        "id": item.pk,
        "os_imovel_id": item.os_imovel_id,
        "imovel_id": item.os_imovel.imovel_id,
        "identificacao": _formatar_identificacao_vinculo(item),
        "endereco": _formatar_endereco_vinculo(item),
        "area_territorial": _formatar_area_vinculo(item),
    }


def _contexto_imoveis_producao(producao):
    vinculados_ids = ProducaoImovel.objects.filter(producao=producao).values_list(
        "os_imovel_id",
        flat=True,
    )
    imoveis_disponiveis = [
        _serializar_os_imovel(vinculo)
        for vinculo in OsImovel.objects.filter(os=producao.os)
        .exclude(pk__in=vinculados_ids)
        .select_related("imovel")
        .order_by("imovel__inscricao_cadastral", "imovel__codigo_isic", "imovel_id")
    ]
    imoveis_producao = [
        _serializar_producao_imovel(item)
        for item in ProducaoImovel.objects.filter(producao=producao)
        .select_related("os_imovel", "os_imovel__imovel")
        .order_by("pk")
    ]
    return imoveis_disponiveis, imoveis_producao


def _validar_justificativa_homologada(producao, justificativa):
    if producao.status == Producao.STATUS_ENVIADO and not (justificativa or "").strip():
        return "Justificativa obrigatória para alterações em produção homologada."
    return None


def _registrar_auditoria_producao_imovel(
    servidor,
    producao_imovel,
    justificativa,
    *,
    operacao="EDICAO_POS_HOMOLOGACAO",
    campo_alterado=None,
    valor_anterior=None,
    valor_novo=None,
):
    LogAuditoria.objects.create(
        servidor=servidor,
        entidade="ProducaoImovel",
        entidade_id=producao_imovel.pk,
        operacao=operacao,
        campo_alterado=campo_alterado,
        valor_anterior=valor_anterior,
        valor_novo=valor_novo,
        justificativa=justificativa,
    )


class ProducaoDetailView(RequerLoginMixin, DetailView):
    model = Producao
    template_name = "producao_detail.html"
    context_object_name = "producao"

    def get_queryset(self):
        return Producao.objects.select_related(
            "os",
            "tipo_producao",
            "criado_por",
            "unidade",
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        producao = self.object
        imoveis = (
            ProducaoImovel.objects.filter(producao=producao)
            .select_related("os_imovel", "os_imovel__imovel")
            .order_by("pk")
        )

        context["os"] = producao.os
        context["imoveis"] = imoveis
        context["status_logs"] = (
            ProducaoStatusLog.objects.filter(producao=producao)
            .select_related(
                "servidor_origem",
                "servidor_destino",
                "unidade_destino",
            )
            .order_by("-data_hora")
        )
        context["tem_servidor"] = _obter_servidor(self.request.user) is not None
        perfil = getattr(self.request, "perfil_acesso", None)
        context["pode_homologar"] = perfil is not None and perfil.pode_homologar
        context["pode_cancelar"] = (
            producao.status != Producao.STATUS_CANCELADO
            and os_editavel_para_usuario(producao.os, self.request)
        )
        context["comentarios"] = (
            Comentario.objects.filter(producao=producao)
            .select_related("servidor")
            .order_by("-data_hora")
        )
        return context


def _verificar_conflito_producao(os_imovel, producao):
    return ProducaoImovel.objects.filter(
        os_imovel__imovel=os_imovel.imovel,
        producao__tipo_producao=producao.tipo_producao,
    ).exclude(
        producao__status__in=[Producao.STATUS_ENVIADO, Producao.STATUS_CANCELADO],
    ).exclude(
        producao=producao,
    ).exists()


def _request_wants_json(request):
    accept = request.headers.get("Accept", "")
    if "application/json" in accept:
        return True
    return request.headers.get("X-Requested-With") == "XMLHttpRequest"


def _serializar_status_log(log):
    return {
        "data_hora": timezone.localtime(log.data_hora).strftime("%d/%m/%Y %H:%M"),
        "status_anterior": log.status_anterior,
        "status_anterior_label": (
            dict(Producao.STATUS_CHOICES).get(log.status_anterior, log.status_anterior)
            if log.status_anterior
            else None
        ),
        "status_novo": log.status_novo,
        "status_novo_label": dict(Producao.STATUS_CHOICES).get(
            log.status_novo,
            log.status_novo,
        ),
        "servidor_origem": log.servidor_origem.nome if log.servidor_origem else None,
        "servidor_destino": log.servidor_destino.nome if log.servidor_destino else None,
        "justificativa": log.justificativa or "",
    }


def _serializar_comentario(comentario):
    badge = "OS"
    if comentario.origem == "PRODUCAO" and comentario.producao:
        badge = comentario.producao.tipo_producao.label_display
    return {
        "id": comentario.pk,
        "texto": comentario.texto,
        "servidor": comentario.servidor.nome,
        "data_hora": timezone.localtime(comentario.data_hora).strftime(
            "%d/%m/%Y %H:%M",
        ),
        "origem": comentario.origem,
        "badge": badge,
    }


def _buscar_registros_siat(busca):
    """Busca registros SIAT por bloco, inscrição ou logradouro."""
    if not SIAT_ARQUIVO_PATH.exists():
        return None, []

    if siat_index.indice_pronto():
        if busca.isdigit() and len(busca) == 12:
            return "bloco", siat_index.buscar_por_bloco(busca)
        if busca.isdigit():
            dados = siat_index.buscar_por_inscricao(int(busca))
            return "inscricao", [dados] if dados else []
        if siat_index.indice_logradouro_disponivel():
            return "logradouro", siat_index.buscar_por_logradouro(busca)
        return "logradouro", buscar_por_logradouro_no_arquivo(
            busca,
            SIAT_ARQUIVO_PATH,
            limite=20,
        )

    if busca.isdigit() and len(busca) == 12:
        registros = buscar_bloco_no_arquivo(busca, SIAT_ARQUIVO_PATH, limite=20)
        return "bloco", registros or []

    if busca.isdigit():
        dados = buscar_inscricao_no_arquivo(int(busca), SIAT_ARQUIVO_PATH)
        return "inscricao", [dados] if dados else []

    return "logradouro", buscar_por_logradouro_no_arquivo(
        busca,
        SIAT_ARQUIVO_PATH,
        limite=20,
    )


def _criar_producao_status_log(
    producao,
    status_anterior,
    status_novo,
    servidor_origem,
    *,
    servidor_destino=None,
    unidade_destino=None,
    justificativa=None,
):
    ProducaoStatusLog.objects.create(
        producao=producao,
        status_anterior=status_anterior,
        status_novo=status_novo,
        servidor_origem=servidor_origem,
        servidor_destino=servidor_destino,
        unidade_destino=unidade_destino,
        justificativa=justificativa or None,
    )


CAMPOS_EDITAVEIS_PRODUCAO = frozenset(
    {
        "numero_producao",
        "numero_sei",
        "observacao",
    },
)

CAMPOS_EDITAVEIS_OS = frozenset(
    {
        "apelido",
        "prioridade",
        "prazo_tipo",
        "prazo_data",
        "data_entrada_divisao",
        "data_entrada_notificacao",
    },
)


def _formatar_data_resposta_json(data):
    if data is None:
        return {"valor": "", "valor_display": "—"}
    return {
        "valor": data.isoformat(),
        "valor_display": data.strftime("%d/%m/%Y"),
    }


class ProducaoEditarCampoView(RequerHomologarMixin, View):
    def post(self, request, pk):
        producao = get_object_or_404(Producao, pk=pk)
        campo = (request.POST.get("campo") or "").strip()
        valor = (request.POST.get("valor") or "").strip()

        if campo not in CAMPOS_EDITAVEIS_PRODUCAO:
            return JsonResponse(
                {"sucesso": False, "erro": "Campo não permitido."},
                status=400,
            )

        if campo in ("numero_producao", "numero_sei", "observacao"):
            setattr(producao, campo, valor or None)
            producao.save(update_fields=[campo])
            return JsonResponse(
                {
                    "sucesso": True,
                    "campo": campo,
                    "valor": getattr(producao, campo) or "",
                },
            )

        return JsonResponse(
            {"sucesso": False, "erro": "Campo não permitido."},
            status=400,
        )


class OSEditarCampoView(RequerLoginMixin, View):
    def post(self, request, pk):
        os_obj = get_object_or_404(
            OS.objects.select_related("tipo_demanda", "finalidade"),
            pk=pk,
        )
        campo = (request.POST.get("campo") or "").strip()
        valor = (request.POST.get("valor") or "").strip()
        perfil = getattr(request, "perfil_acesso", None)

        if campo not in CAMPOS_EDITAVEIS_OS:
            return JsonResponse(
                {"sucesso": False, "erro": "Campo não permitido."},
                status=400,
            )

        if campo == "data_entrada_divisao":
            if not _pode_editar_entrada_dai(request):
                return JsonResponse(
                    {"sucesso": False, "erro": MSG_SEM_PERMISSAO},
                    status=403,
                )
            if not valor:
                data_valor = None
            else:
                try:
                    data_valor = datetime.date.fromisoformat(valor)
                except ValueError:
                    return JsonResponse(
                        {"sucesso": False, "erro": "Data inválida."},
                        status=400,
                    )
            os_obj.data_entrada_divisao = data_valor
            os_obj.save(update_fields=["data_entrada_divisao"])
            vinculo = OsProcesso.objects.filter(
                os=os_obj,
                tipo_vinculo="PRINCIPAL",
            ).first()
            if vinculo:
                vinculo.data_entrada_divisao = data_valor
                vinculo.save(update_fields=["data_entrada_divisao"])
            return JsonResponse(
                {
                    "sucesso": True,
                    "campo": campo,
                    **_formatar_data_resposta_json(data_valor),
                },
            )

        if campo == "data_entrada_notificacao":
            if not os_editavel_para_usuario(os_obj, request):
                return JsonResponse(
                    {"sucesso": False, "erro": MSG_OS_SOMENTE_LEITURA},
                    status=403,
                )
            if not valor:
                return JsonResponse(
                    {"sucesso": False, "erro": "Informe a data."},
                    status=400,
                )
            try:
                data_valor = datetime.date.fromisoformat(valor)
            except ValueError:
                return JsonResponse(
                    {"sucesso": False, "erro": "Data inválida."},
                    status=400,
                )
            os_obj.data_entrada_notificacao = data_valor
            os_obj.save(update_fields=["data_entrada_notificacao"])
            return JsonResponse(
                {
                    "sucesso": True,
                    "campo": campo,
                    **_formatar_data_resposta_json(data_valor),
                },
            )

        if not perfil or not perfil.pode_homologar:
            return JsonResponse(
                {"sucesso": False, "erro": MSG_SEM_PERMISSAO},
                status=403,
            )

        if campo == "apelido":
            os_obj.apelido = valor or None
            os_obj.save(update_fields=["apelido"])
            return JsonResponse(
                {"sucesso": True, "campo": campo, "valor": os_obj.apelido or ""},
            )

        if campo == "prioridade":
            if valor not in {"NORMAL", "PRIORITARIO", "URGENTE"}:
                return JsonResponse(
                    {"sucesso": False, "erro": "Prioridade inválida."},
                    status=400,
                )
            os_obj.prioridade = valor
            os_obj.save(update_fields=["prioridade"])
            return JsonResponse({"sucesso": True, "campo": campo, "valor": valor})

        if campo == "prazo_tipo":
            tipos = {item[0] for item in OS.PRAZO_TIPO_CHOICES}
            if valor not in tipos:
                return JsonResponse(
                    {"sucesso": False, "erro": "Tipo de prazo inválido."},
                    status=400,
                )
            os_obj.prazo_tipo = valor
            os_obj.save(update_fields=["prazo_tipo"])
            return JsonResponse(
                {
                    "sucesso": True,
                    "campo": campo,
                    "valor": dict(OS.PRAZO_TIPO_CHOICES).get(valor, valor),
                },
            )

        if campo == "prazo_data":
            if not valor:
                os_obj.prazo_data = None
            else:
                try:
                    os_obj.prazo_data = datetime.date.fromisoformat(valor)
                except ValueError:
                    return JsonResponse(
                        {"sucesso": False, "erro": "Data inválida."},
                        status=400,
                    )
            os_obj.save(update_fields=["prazo_data"])
            hoje = timezone.localdate()
            dias = (
                (os_obj.prazo_data - hoje).days if os_obj.prazo_data else None
            )
            resposta = _formatar_data_resposta_json(os_obj.prazo_data)
            resposta["dias_sei"] = dias
            return JsonResponse({"sucesso": True, "campo": campo, **resposta})

        return JsonResponse(
            {"sucesso": False, "erro": "Campo não suportado."},
            status=400,
        )


class PreferenciaGerencialView(RequerLoginJSONMixin, View):
    def get(self, request):
        servidor = _obter_servidor(request.user)
        if servidor is None:
            return JsonResponse({"error": MSG_SEM_PERMISSAO}, status=403)
        return JsonResponse(
            {
                "colunas_visiveis": _colunas_visiveis_gerencial(servidor),
                "colunas_disponiveis": [
                    {"id": chave, "label": meta["label"]}
                    for chave, meta in COLUNAS_GERENCIAL_CONFIG.items()
                ],
            },
        )

    def post(self, request):
        servidor = _obter_servidor(request.user)
        if servidor is None:
            return JsonResponse({"sucesso": False, "erro": MSG_SEM_PERMISSAO}, status=403)

        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return JsonResponse(
                {"sucesso": False, "erro": "JSON inválido."},
                status=400,
            )

        colunas = payload.get("colunas_visiveis", [])
        if not isinstance(colunas, list):
            return JsonResponse(
                {"sucesso": False, "erro": "colunas_visiveis deve ser uma lista."},
                status=400,
            )

        validas = [c for c in colunas if c in COLUNAS_GERENCIAL_CONFIG]
        preferencia, _ = PreferenciaGerencial.objects.get_or_create(servidor=servidor)
        preferencia.colunas_visiveis = validas or list(COLUNAS_GERENCIAL_PADRAO)
        preferencia.save(update_fields=["colunas_visiveis"])
        return JsonResponse(
            {
                "sucesso": True,
                "colunas_visiveis": preferencia.colunas_visiveis,
            },
        )


class ComentarioCreateView(RequerLoginJSONMixin, View):
    def post(self, request, os_pk=None, prod_pk=None):
        servidor = _obter_servidor(request.user)
        if servidor is None:
            return JsonResponse(
                {"sucesso": False, "erro": MSG_SEM_PERMISSAO},
                status=403,
            )

        texto = (request.POST.get("texto") or "").strip()
        if not texto:
            return JsonResponse(
                {"sucesso": False, "erro": "Texto do comentário é obrigatório."},
                status=400,
            )

        if prod_pk is not None:
            producao = get_object_or_404(Producao.objects.select_related("tipo_producao"), pk=prod_pk)
            comentario = Comentario.objects.create(
                os=producao.os,
                producao=producao,
                origem="PRODUCAO",
                texto=texto,
                servidor=servidor,
            )
            comentario = Comentario.objects.select_related(
                "servidor",
                "producao",
                "producao__tipo_producao",
            ).get(pk=comentario.pk)
        else:
            os_obj = get_object_or_404(OS, pk=os_pk)
            comentario = Comentario.objects.create(
                os=os_obj,
                origem="OS",
                texto=texto,
                servidor=servidor,
            )
            comentario = Comentario.objects.select_related("servidor").get(pk=comentario.pk)

        return JsonResponse(
            {
                "sucesso": True,
                "comentario": _serializar_comentario(comentario),
            },
        )


def _parse_json_ou_post(request):
    if request.content_type and "application/json" in request.content_type:
        try:
            return json.loads(request.body.decode("utf-8") or "{}")
        except (json.JSONDecodeError, UnicodeDecodeError):
            return {}
    return request.POST


def _opcoes_pos_enviado(os_obj):
    return [
        {"label": "Manter em atendimento", "acao": "manter"},
        {
            "label": "Encaminhar para unidade",
            "acao": "encaminhar",
            "url": reverse("os_encaminhar", kwargs={"pk": os_obj.pk}),
        },
        {
            "label": "Encerrar na Divisão",
            "acao": "encerrar",
            "url": reverse("os_encerrar", kwargs={"pk": os_obj.pk}),
        },
    ]


class OSEtapaAPIView(RequerLoginJSONMixin, View):
    def post(self, request, pk):
        from core.os_service import ETAPAS_INTERNAS_LABELS

        os_obj = get_object_or_404(OS, pk=pk)
        servidor = _obter_servidor(request.user)
        if servidor is None:
            return JsonResponse(
                {"sucesso": False, "erro": MSG_SEM_PERMISSAO},
                status=403,
            )
        if not os_editavel_para_usuario(os_obj, request):
            return JsonResponse(
                {"sucesso": False, "erro": MSG_OS_SOMENTE_LEITURA},
                status=403,
            )

        dados = _parse_json_ou_post(request)
        etapa = (dados.get("etapa_interna") or "").strip()
        if etapa not in ETAPAS_FLUXO_GERENCIAL and etapa != "CONCLUIDA":
            return JsonResponse(
                {"sucesso": False, "erro": "Etapa interna inválida."},
                status=400,
            )

        unidade = _obter_unidade_principal_servidor(servidor)
        if unidade is None:
            return JsonResponse(
                {"sucesso": False, "erro": "Unidade do servidor não encontrada."},
                status=400,
            )

        tarefa = (
            TarefaInterna.objects.filter(
                os=os_obj,
                unidade=unidade,
                status="PENDENTE",
            )
            .order_by("-data_inicio")
            .first()
        )
        etapa_atual = tarefa.etapa_interna if tarefa else None
        permitidas = {
            item["valor"] for item in _etapas_posteriores_gerencial(etapa_atual)
        }
        if etapa not in permitidas:
            return JsonResponse(
                {"sucesso": False, "erro": "Etapa não disponível a partir da atual."},
                status=400,
            )

        with transaction.atomic():
            encaminhamento = Encaminhamento.objects.create(
                os=os_obj,
                unidade_interna_origem=unidade,
                servidor_origem=servidor,
                unidade_interna_destino=unidade,
                servidor_destino=None,
                etapa_interna=etapa,
                tipo_macroetapa=None,
                aguarda_retorno=False,
                automatico=True,
                observacao=f"Etapa interna alterada para {etapa} via gerencial.",
                manter_aberta_na_unidade=True,
            )
            if tarefa:
                tarefa.etapa_interna = etapa
                tarefa.save(update_fields=["etapa_interna"])
            else:
                TarefaInterna.objects.create(
                    os=os_obj,
                    encaminhamento=encaminhamento,
                    unidade=unidade,
                    servidor=servidor,
                    etapa_interna=etapa,
                    status="PENDENTE",
                    data_inicio=timezone.now(),
                )

        return JsonResponse(
            {
                "sucesso": True,
                "etapa_label": ETAPAS_INTERNAS_LABELS.get(etapa, etapa),
                "etapa_interna": etapa,
            },
        )


class ProducaoStatusAPIView(RequerLoginJSONMixin, View):
    """Endpoint de cancelamento de produção.

    Produção é um registro concluído no momento da criação; a única transição
    de status disponível é o cancelamento (ENVIADO -> CANCELADO).
    """

    def post(self, request, pk):
        producao = get_object_or_404(
            Producao.objects.select_related("os", "tipo_producao"),
            pk=pk,
        )
        servidor = _obter_servidor(request.user)
        if servidor is None:
            return JsonResponse(
                {"sucesso": False, "erro": MSG_SEM_PERMISSAO},
                status=403,
            )
        if not os_editavel_para_usuario(producao.os, request):
            return JsonResponse(
                {"sucesso": False, "erro": MSG_OS_SOMENTE_LEITURA},
                status=403,
            )

        dados = _parse_json_ou_post(request)
        status_novo = (dados.get("status") or "").strip()
        if status_novo != Producao.STATUS_CANCELADO:
            return JsonResponse(
                {"sucesso": False, "erro": "Somente o cancelamento é permitido."},
                status=400,
            )
        if producao.status == Producao.STATUS_CANCELADO:
            return JsonResponse(
                {"sucesso": False, "erro": "Produção já está cancelada."},
                status=400,
            )

        justificativa = (dados.get("justificativa") or "").strip() or None
        status_anterior = producao.status
        producao.status = Producao.STATUS_CANCELADO
        producao.save(update_fields=["status"])

        _criar_producao_status_log(
            producao,
            status_anterior,
            Producao.STATUS_CANCELADO,
            servidor,
            justificativa=justificativa,
        )

        LogAuditoria.objects.create(
            servidor=servidor,
            entidade="Producao",
            entidade_id=producao.pk,
            operacao="CANCELAMENTO",
            campo_alterado="status",
            valor_anterior=status_anterior,
            valor_novo=Producao.STATUS_CANCELADO,
            justificativa=justificativa,
        )

        return JsonResponse(
            {
                "sucesso": True,
                "status": producao.status,
                "status_label": _label_status_producao_gerencial(producao.status),
                "cor": _cor_status_producao_gerencial(producao.status),
            },
        )


class OSComentariosAPIView(RequerLoginJSONMixin, View):
    def _filtrar_comentarios(self, os_obj, request):
        qs = (
            Comentario.objects.filter(os=os_obj)
            .select_related("servidor", "producao", "producao__tipo_producao")
            .order_by("-data_hora")
        )
        producao_id = (request.GET.get("producao") or "").strip()
        if not producao_id and request.method == "POST":
            dados = _parse_json_ou_post(request)
            producao_id = (dados.get("producao") or dados.get("producao_id") or "").strip()
        if producao_id:
            try:
                qs = qs.filter(producao_id=int(producao_id))
            except (ValueError, TypeError):
                qs = qs.none()
        else:
            qs = qs.filter(origem="OS")
        return qs

    def get(self, request, pk):
        os_obj = get_object_or_404(OS, pk=pk)
        qs = self._filtrar_comentarios(os_obj, request)
        total = qs.count()
        comentarios = [_serializar_comentario(c) for c in qs[:5]]
        return JsonResponse({"comentarios": comentarios, "total": total})

    def post(self, request, pk):
        os_obj = get_object_or_404(OS, pk=pk)
        servidor = _obter_servidor(request.user)
        if servidor is None:
            return JsonResponse(
                {"sucesso": False, "erro": MSG_SEM_PERMISSAO},
                status=403,
            )
        dados = _parse_json_ou_post(request)
        texto = (dados.get("texto") or "").strip()
        if not texto:
            return JsonResponse(
                {"sucesso": False, "erro": "Texto do comentário é obrigatório."},
                status=400,
            )
        producao_id = dados.get("producao") or dados.get("producao_id")
        producao = None
        if producao_id:
            producao = get_object_or_404(Producao, pk=int(producao_id), os=os_obj)
        comentario = Comentario.objects.create(
            os=os_obj,
            producao=producao,
            origem="PRODUCAO" if producao else "OS",
            texto=texto,
            servidor=servidor,
        )
        comentario = Comentario.objects.select_related(
            "servidor",
            "producao",
            "producao__tipo_producao",
        ).get(pk=comentario.pk)
        qs = self._filtrar_comentarios(os_obj, request)
        return JsonResponse(
            {
                "sucesso": True,
                "comentario": _serializar_comentario(comentario),
                "comentarios": [
                    _serializar_comentario(c) for c in qs[:5]
                ],
                "total": qs.count(),
            },
        )


class ProducaoNovaAPIView(RequerLoginJSONMixin, View):
    def post(self, request, pk):
        os_obj = get_object_or_404(OS, pk=pk)
        servidor = _obter_servidor(request.user)
        if servidor is None:
            return JsonResponse(
                {"sucesso": False, "erro": MSG_SEM_PERMISSAO},
                status=403,
            )
        if not os_editavel_para_usuario(os_obj, request):
            return JsonResponse(
                {"sucesso": False, "erro": MSG_OS_SOMENTE_LEITURA},
                status=403,
            )
        if getattr(request, "visibilidade", "UNIDADE") == "DEPARTAMENTO":
            return JsonResponse(
                {
                    "sucesso": False,
                    "erro": "Perfil DEPARTAMENTO não pode registrar produção.",
                },
                status=403,
            )

        unidade = _obter_unidade_principal_servidor(servidor)
        dados = _parse_json_ou_post(request)
        tipo_id = dados.get("tipo_producao") or dados.get("tipo_producao_id")
        if not tipo_id:
            return JsonResponse(
                {"sucesso": False, "erro": "Selecione o tipo de produção."},
                status=400,
            )
        try:
            tipo_qs = TipoProducao.objects.filter(ativo=True)
            if unidade:
                tipo_qs = tipo_qs.filter(
                    unidades_competentes__unidade_interna=unidade,
                )
            tipo_producao = tipo_qs.distinct().get(pk=int(tipo_id))
        except (TipoProducao.DoesNotExist, ValueError, TypeError):
            return JsonResponse(
                {"sucesso": False, "erro": "Tipo de produção inválido."},
                status=400,
            )

        observacao = (dados.get("observacao") or "").strip() or None
        numero_producao = None
        if tipo_producao.prefixo != "Despacho":
            numero_producao = _gerar_numero_producao(tipo_producao)
        producao = Producao.objects.create(
            os=os_obj,
            tipo_producao=tipo_producao,
            numero_producao=numero_producao,
            ano=timezone.localdate().year,
            status=Producao.STATUS_ENVIADO,
            data_enviado=timezone.localdate(),
            unidade=unidade,
            criado_por=servidor,
            observacao=observacao,
        )
        _criar_producao_status_log(
            producao,
            None,
            Producao.STATUS_ENVIADO,
            servidor,
        )
        ativar_atendimento_interno_se_necessario(os_obj, servidor=servidor)
        if unidade:
            registrar_em_atendimento_na_unidade(os_obj, unidade, servidor=servidor)

        return JsonResponse(
            {
                "sucesso": True,
                "producao_pk": producao.pk,
                "status_label": _label_status_producao_gerencial(producao.status),
            },
        )


class ProducaoVincularImovelView(RequerLoginMixin, View):
    template_name = "producao_vincular_imovel.html"

    def dispatch(self, request, *args, **kwargs):
        self.producao_obj = get_object_or_404(
            Producao.objects.select_related("os", "tipo_producao"),
            pk=kwargs["pk"],
        )
        if _obter_servidor(request.user) is None:
            if request.content_type and "application/json" in request.content_type:
                return JsonResponse(
                    {"sucesso": False, "erro": MSG_SEM_PERMISSAO},
                    status=403,
                )
            messages.error(request, MSG_SEM_PERMISSAO)
            return redirect(reverse("producao_detail", kwargs={"pk": self.producao_obj.pk}))
        return super().dispatch(request, *args, **kwargs)

    @method_decorator(ensure_csrf_cookie)
    def get(self, request, *args, **kwargs):
        imoveis_disponiveis, imoveis_producao = _contexto_imoveis_producao(self.producao_obj)
        return render(
            request,
            self.template_name,
            {
                "producao": self.producao_obj,
                "os": self.producao_obj.os,
                "imoveis_disponiveis": imoveis_disponiveis,
                "imoveis_producao": imoveis_producao,
                "producao_homologada": self.producao_obj.status == Producao.STATUS_ENVIADO,
            },
        )

    def _resposta_json(self, request, *, sucesso=True, status=200, **dados):
        imoveis_disponiveis, imoveis_producao = _contexto_imoveis_producao(self.producao_obj)
        payload = {
            "sucesso": sucesso,
            "imoveis_disponiveis": imoveis_disponiveis,
            "imoveis_producao": imoveis_producao,
        }
        payload.update(dados)
        return JsonResponse(payload, status=status)

    def _obter_producao_imovel(self, producao_imovel_id):
        if producao_imovel_id in (None, ""):
            return None
        try:
            producao_imovel_id = int(producao_imovel_id)
        except (TypeError, ValueError):
            return None
        return ProducaoImovel.objects.filter(
            pk=producao_imovel_id,
            producao=self.producao_obj,
        ).first()

    def _auditar_se_homologada(self, servidor, producao_imovel, justificativa, **kwargs):
        if self.producao_obj.status == Producao.STATUS_ENVIADO:
            _registrar_auditoria_producao_imovel(
                servidor,
                producao_imovel,
                justificativa,
                **kwargs,
            )

    def _processar_acao(self, request, acao, servidor, justificativa):
        tipo = acao.get("tipo")

        if tipo == "vincular":
            os_imovel_id = acao.get("os_imovel_id")
            if not os_imovel_id:
                return {"sucesso": False, "erro": "os_imovel_id é obrigatório."}

            os_imovel = OsImovel.objects.filter(
                pk=os_imovel_id,
                os=self.producao_obj.os,
            ).first()
            if os_imovel is None:
                return {"sucesso": False, "erro": "Imóvel não pertence à OS desta produção."}

            if ProducaoImovel.objects.filter(
                producao=self.producao_obj,
                os_imovel=os_imovel,
            ).exists():
                return {"sucesso": False, "erro": "Imóvel já vinculado à produção."}

            if _verificar_conflito_producao(os_imovel, self.producao_obj):
                prefixo = self.producao_obj.tipo_producao.prefixo
                return {
                    "sucesso": False,
                    "erro": (
                        f"Esta inscrição já está em outra produção ativa do tipo {prefixo}. "
                        "Homologue ou cancele a produção anterior antes de incluir aqui."
                    ),
                }

            producao_imovel = ProducaoImovel.objects.create(
                producao=self.producao_obj,
                os_imovel=os_imovel,
            )
            self._auditar_se_homologada(
                servidor,
                producao_imovel,
                justificativa,
                operacao="EDICAO_POS_HOMOLOGACAO",
                campo_alterado="vinculo",
                valor_novo=str(os_imovel.pk),
            )
            return {
                "sucesso": True,
                "tipo": tipo,
                "item": _serializar_producao_imovel(producao_imovel),
            }

        if tipo == "desvincular":
            producao_imovel = self._obter_producao_imovel(acao.get("producao_imovel_id"))
            if producao_imovel is None:
                return {"sucesso": False, "erro": "Imóvel da produção não encontrado."}
            item_id = producao_imovel.pk
            os_imovel_id = producao_imovel.os_imovel_id
            self._auditar_se_homologada(
                servidor,
                producao_imovel,
                justificativa,
                operacao="EDICAO_POS_HOMOLOGACAO",
                campo_alterado="vinculo",
                valor_anterior=str(os_imovel_id),
                valor_novo=None,
            )
            producao_imovel.delete()
            return {
                "sucesso": True,
                "tipo": tipo,
                "id": item_id,
                "os_imovel_id": os_imovel_id,
            }

        return {"sucesso": False, "erro": f"Ação desconhecida: {tipo}"}

    def post(self, request, *args, **kwargs):
        servidor = _obter_servidor(request.user)
        if servidor is None:
            return JsonResponse(
                {"sucesso": False, "erro": MSG_SEM_PERMISSAO},
                status=403,
            )

        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return JsonResponse({"sucesso": False, "erro": "JSON inválido."}, status=400)

        acoes = payload.get("acoes") or []
        if not acoes:
            return JsonResponse({"sucesso": False, "erro": "Nenhuma ação informada."}, status=400)

        logger.info(
            "ProducaoVincularImovelView POST producao=%s acoes=%s",
            self.producao_obj.pk,
            acoes,
        )

        justificativa = (payload.get("justificativa") or "").strip()
        erro_justificativa = _validar_justificativa_homologada(
            self.producao_obj,
            justificativa,
        )
        if erro_justificativa:
            return JsonResponse({"sucesso": False, "erro": erro_justificativa}, status=400)

        resultados = []
        with transaction.atomic():
            for acao in acoes:
                resultado = self._processar_acao(request, acao, servidor, justificativa)
                if not resultado.get("sucesso"):
                    status = resultado.pop("status", 400)
                    return self._resposta_json(
                        request,
                        sucesso=False,
                        status=status,
                        erro=resultado.get("erro", "Erro ao processar ação."),
                        resultados=resultados,
                    )
                resultados.append(resultado)

        return self._resposta_json(
            request,
            sucesso=True,
            mensagem="Alterações salvas com sucesso.",
            resultados=resultados,
        )


class OSVincularImovelView(RequerLoginMixin, View):
    template_name = "os_vincular_imovel.html"

    def dispatch(self, request, *args, **kwargs):
        self.os_obj = get_object_or_404(OS, pk=kwargs["pk"])
        if not os_editavel_para_usuario(self.os_obj, request):
            messages.error(request, MSG_OS_SOMENTE_LEITURA)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, {"os": self.os_obj})

    def post(self, request, *args, **kwargs):
        servidor = _obter_servidor(request.user)
        if servidor is None:
            messages.error(request, MSG_SEM_PERMISSAO)
            return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))

        dados_json = request.POST.get("dados_completos", "").strip()
        if not dados_json:
            messages.error(request, "Dados do imóvel não informados.")
            return redirect(reverse("os_vincular_imovel", kwargs={"pk": self.os_obj.pk}))

        try:
            dados_completos = json.loads(dados_json)
        except json.JSONDecodeError:
            messages.error(request, "Dados do imóvel inválidos.")
            return redirect(reverse("os_vincular_imovel", kwargs={"pk": self.os_obj.pk}))

        if dados_completos.get("codigo_isic"):
            imovel = Imovel.objects.filter(
                codigo_isic=dados_completos.get("codigo_isic"),
            ).first()
            if imovel and OsImovel.objects.filter(
                os=self.os_obj,
                imovel=imovel,
            ).exists():
                messages.error(request, "Este imóvel já está vinculado à OS.")
                return redirect(
                    reverse("os_vincular_imovel", kwargs={"pk": self.os_obj.pk}),
                )
            vincular_isic_a_os(self.os_obj, dados_completos, servidor)
        else:
            inscricao = dados_completos.get("inscricao_cadastral")
            imovel = Imovel.objects.filter(inscricao_cadastral=inscricao).first()
            if imovel and OsImovel.objects.filter(
                os=self.os_obj,
                imovel=imovel,
            ).exists():
                messages.error(request, "Este imóvel já está vinculado à OS.")
                return redirect(
                    reverse("os_vincular_imovel", kwargs={"pk": self.os_obj.pk}),
                )
            vincular_imovel_a_os(self.os_obj, dados_completos, servidor)
        messages.success(request, "Imóvel vinculado à OS com sucesso.")
        return redirect(reverse("os_detalhe", kwargs={"pk": self.os_obj.pk}))


def _adicionar_resultado_siat(resultados, vistos, dados):
    chave = f"siat:{dados.get('inscricao_cadastral')}"
    if chave in vistos:
        return False
    vistos.add(chave)
    resultados.append(_montar_resultado_busca(dados, "siat"))
    return True


class BuscarImoveisAPIView(RequerLoginJSONMixin, View):
    def get(self, request):
        busca = request.GET.get("q", "").strip()
        if not busca:
            return JsonResponse([], safe=False)

        resultados = []
        vistos = set()

        if SIAT_ARQUIVO_PATH.exists():
            if busca.isdigit() and len(busca) == 12:
                _, registros = _buscar_registros_siat(busca)
                for dados in registros:
                    _adicionar_resultado_siat(resultados, vistos, dados)
                return JsonResponse(resultados[:20], safe=False)

            if busca.isdigit():
                _, registros = _buscar_registros_siat(busca)
                for dados in registros:
                    _adicionar_resultado_siat(resultados, vistos, dados)
                return JsonResponse(resultados[:20], safe=False)

        for imovel in Imovel.objects.filter(
            tipo_identificacao="ISIC",
            codigo_isic__icontains=busca,
        ).order_by("codigo_isic")[:20]:
            dados = _os_imovel_para_dict(_ultimo_os_imovel(imovel))
            dados["codigo_isic"] = imovel.codigo_isic
            chave = f"isic:{imovel.pk}"
            if chave not in vistos:
                vistos.add(chave)
                resultados.append(_montar_resultado_busca(dados, "isic"))

        if SIAT_ARQUIVO_PATH.exists():
            _, registros = _buscar_registros_siat(busca)
            for dados in registros:
                if _adicionar_resultado_siat(resultados, vistos, dados):
                    if len(resultados) >= 20:
                        break

        return JsonResponse(resultados[:20], safe=False)


class SiatCoordenadasBlocoView(RequerLoginJSONMixin, View):
    def get(self, request, num_bloco):
        if not SIAT_ARQUIVO_PATH.exists():
            return JsonResponse({"encontrado": False})

        coordenadas = obter_coordenadas_bloco(num_bloco, SIAT_ARQUIVO_PATH)
        if not coordenadas:
            return JsonResponse({"encontrado": False})

        return JsonResponse(
            {
                "encontrado": True,
                "latitude": _decimal_para_json(coordenadas.get("latitude")),
                "longitude": _decimal_para_json(coordenadas.get("longitude")),
                "coord_x": _decimal_para_json(coordenadas.get("coord_x")),
                "coord_y": _decimal_para_json(coordenadas.get("coord_y")),
            },
        )
