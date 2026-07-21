from datetime import date

import openpyxl
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import OsProcesso, Producao


def _dias_tempo_registro(data_vinculo, data_entrada_divisao):
    if not data_vinculo or not data_entrada_divisao:
        return None
    if timezone.is_aware(data_vinculo):
        data_registro = timezone.localtime(data_vinculo).date()
    else:
        data_registro = data_vinculo.date()
    return (data_registro - data_entrada_divisao).days


def relatorio_tempo_registro_processo(filtros):
    qs = (
        OsProcesso.objects.filter(
            data_entrada_divisao__isnull=False,
            data_vinculo__isnull=False,
        )
        .select_related("os", "processo_sei", "os__criado_por")
        .order_by("data_vinculo")
    )

    if filtros.get("data_inicio"):
        qs = qs.filter(data_vinculo__date__gte=filtros["data_inicio"])
    if filtros.get("data_fim"):
        qs = qs.filter(data_vinculo__date__lte=filtros["data_fim"])
    if filtros.get("servidor_id"):
        qs = qs.filter(os__criado_por__id=filtros["servidor_id"])

    return qs


def linhas_relatorio_tempo_registro(queryset):
    linhas = []
    for vinculo in queryset:
        dias = _dias_tempo_registro(vinculo.data_vinculo, vinculo.data_entrada_divisao)
        linhas.append(
            {
                "numero_os": vinculo.os.numero_os,
                "processo_sei": vinculo.processo_sei.numero_processo,
                "criado_por": (
                    vinculo.os.criado_por.nome if vinculo.os.criado_por else "—"
                ),
                "data_entrada_divisao": vinculo.data_entrada_divisao,
                "data_vinculo": vinculo.data_vinculo,
                "dias": dias,
            },
        )
    return linhas


def exportar_tempo_registro_excel(queryset):
    vinculos = list(queryset)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tempo de Registro"

    cabecalho_fill = PatternFill(
        start_color="1A3A5C",
        end_color="1A3A5C",
        fill_type="solid",
    )
    cabecalho_font = Font(color="FFFFFF", bold=True, size=11)

    colunas = [
        ("OS", 18),
        ("Processo SEI", 25),
        ("Criado por", 25),
        ("Entrada na Divisão", 18),
        ("Registro SIPRAC", 20),
        ("Dias de diferença", 16),
    ]

    for col, (titulo, largura) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = cabecalho_font
        cell.fill = cabecalho_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = largura

    for row, vinculo in enumerate(vinculos, 2):
        dias = _dias_tempo_registro(vinculo.data_vinculo, vinculo.data_entrada_divisao)
        ws.cell(row=row, column=1, value=vinculo.os.numero_os)
        ws.cell(row=row, column=2, value=vinculo.processo_sei.numero_processo)
        ws.cell(
            row=row,
            column=3,
            value=vinculo.os.criado_por.nome if vinculo.os.criado_por else "—",
        )
        ws.cell(
            row=row,
            column=4,
            value=(
                vinculo.data_entrada_divisao.strftime("%d/%m/%Y")
                if vinculo.data_entrada_divisao
                else "—"
            ),
        )
        ws.cell(
            row=row,
            column=5,
            value=(
                timezone.localtime(vinculo.data_vinculo).strftime("%d/%m/%Y %H:%M")
                if vinculo.data_vinculo
                else "—"
            ),
        )
        dias_cell = ws.cell(
            row=row,
            column=6,
            value=dias if dias is not None else "—",
        )
        if dias is not None and dias > 5:
            dias_cell.font = Font(color="C00000", bold=True)
        elif dias is not None and dias > 2:
            dias_cell.font = Font(color="ED7D31", bold=True)

        if row % 2 == 0:
            fill = PatternFill(
                start_color="EEF2F7",
                end_color="EEF2F7",
                fill_type="solid",
            )
            for col in range(1, len(colunas) + 1):
                ws.cell(row=row, column=col).fill = fill

    total_row = len(vinculos) + 2
    ws.cell(row=total_row, column=1, value=f"Total: {len(vinculos)} registros")
    ws.cell(row=total_row, column=1).font = Font(bold=True)

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="tempo_registro_processo_{date.today():%Y%m%d}.xlsx"'
    )
    wb.save(response)
    return response


def relatorio_producao_por_servidor(filtros):
    """
    Gera dados do relatório de produção por servidor.
    filtros: dict com servidor_id, data_inicio, data_fim,
             tipo_producao_id, unidade_id (todos opcionais)
    Retorna queryset de Producao com select_related.
    """
    qs = (
        Producao.objects.filter(status=Producao.STATUS_ENVIADO)
        .select_related(
            "tipo_producao",
            "os",
            "os__natureza",
            "autor_trabalho",
            "homologado_por",
        )
        .order_by("autor_trabalho__nome", "data_enviado")
    )

    if filtros.get("servidor_id"):
        qs = qs.filter(autor_trabalho__id=filtros["servidor_id"])
    if filtros.get("data_inicio"):
        qs = qs.filter(data_enviado__gte=filtros["data_inicio"])
    if filtros.get("data_fim"):
        qs = qs.filter(data_enviado__lte=filtros["data_fim"])
    if filtros.get("tipo_producao_id"):
        qs = qs.filter(tipo_producao__id=filtros["tipo_producao_id"])
    if filtros.get("unidade_id"):
        hoje = timezone.localdate()
        qs = qs.filter(
            autor_trabalho__vinculos_unidade__unidade_id=filtros["unidade_id"],
        ).filter(
            Q(autor_trabalho__vinculos_unidade__data_fim__isnull=True)
            | Q(autor_trabalho__vinculos_unidade__data_fim__gte=hoje),
        ).distinct()

    return qs


def linhas_relatorio_producao(queryset):
    """Serializa produções para exibição na prévia HTML."""
    producoes = list(queryset)
    processos_por_os = _mapa_processos_principais(producoes)
    linhas = []
    for producao in producoes:
        processo_principal = processos_por_os.get(producao.os_id)
        linhas.append(
            {
                "autor": (
                    producao.autor_trabalho.nome if producao.autor_trabalho else "—"
                ),
                "tipo": (
                    producao.tipo_producao.prefixo
                    if producao.tipo_producao
                    else "Despacho"
                ),
                "numero_producao": producao.numero_producao or "—",
                "numero_os": producao.os.numero_os,
                "natureza": (
                    producao.os.natureza.descricao if producao.os.natureza else "—"
                ),
                "processo_sei": (
                    processo_principal.processo_sei.numero_processo
                    if processo_principal and processo_principal.processo_sei
                    else "—"
                ),
                "data_enviado": producao.data_enviado,
                "homologado_por": (
                    producao.homologado_por.nome if producao.homologado_por else "—"
                ),
                "unidade": _sigla_unidade_autor(producao.autor_trabalho),
            },
        )
    return linhas


def _mapa_processos_principais(producoes):
    os_ids = {producao.os_id for producao in producoes}
    if not os_ids:
        return {}
    return {
        vinculo.os_id: vinculo
        for vinculo in OsProcesso.objects.filter(
            os_id__in=os_ids,
            tipo_vinculo="PRINCIPAL",
        ).select_related("processo_sei")
    }


def _sigla_unidade_autor(servidor):
    if servidor is None:
        return "—"
    hoje = timezone.localdate()
    vinculo = (
        servidor.vinculos_unidade.filter(
            Q(data_fim__isnull=True) | Q(data_fim__gte=hoje),
        )
        .select_related("unidade")
        .first()
    )
    return vinculo.unidade.sigla if vinculo else "—"


def exportar_producao_excel(queryset):
    """
    Gera arquivo Excel do relatório de produção.
    Retorna HttpResponse com o arquivo.
    """
    producoes = list(queryset)
    processos_por_os = _mapa_processos_principais(producoes)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produção por Servidor"

    cabecalho_fill = PatternFill(
        start_color="1A3A5C",
        end_color="1A3A5C",
        fill_type="solid",
    )
    cabecalho_font = Font(color="FFFFFF", bold=True, size=11)

    colunas = [
        ("Autor do trabalho", 25),
        ("Tipo", 10),
        ("Número produção", 18),
        ("OS vinculada", 18),
        ("Natureza", 20),
        ("Processo SEI", 25),
        ("Data homologação", 18),
        ("Homologado por", 25),
        ("Unidade", 10),
    ]

    for col, (titulo, largura) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = cabecalho_font
        cell.fill = cabecalho_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = largura

    for row, producao in enumerate(producoes, 2):
        processo_principal = processos_por_os.get(producao.os_id)

        ws.cell(
            row=row,
            column=1,
            value=producao.autor_trabalho.nome if producao.autor_trabalho else "—",
        )
        ws.cell(
            row=row,
            column=2,
            value=(
                producao.tipo_producao.prefixo
                if producao.tipo_producao
                else "Despacho"
            ),
        )
        ws.cell(row=row, column=3, value=producao.numero_producao or "—")
        ws.cell(row=row, column=4, value=producao.os.numero_os)
        ws.cell(
            row=row,
            column=5,
            value=producao.os.natureza.descricao if producao.os.natureza else "—",
        )
        ws.cell(
            row=row,
            column=6,
            value=(
                processo_principal.processo_sei.numero_processo
                if processo_principal and processo_principal.processo_sei
                else "—"
            ),
        )
        ws.cell(
            row=row,
            column=7,
            value=(
                producao.data_enviado.strftime("%d/%m/%Y")
                if producao.data_enviado
                else "—"
            ),
        )
        ws.cell(
            row=row,
            column=8,
            value=producao.homologado_por.nome if producao.homologado_por else "—",
        )
        ws.cell(
            row=row,
            column=9,
            value=_sigla_unidade_autor(producao.autor_trabalho),
        )

        if row % 2 == 0:
            fill = PatternFill(
                start_color="EEF2F7",
                end_color="EEF2F7",
                fill_type="solid",
            )
            for col in range(1, len(colunas) + 1):
                ws.cell(row=row, column=col).fill = fill

    total_row = len(producoes) + 2
    ws.cell(row=total_row, column=1, value=f"Total: {len(producoes)} produções")
    ws.cell(row=total_row, column=1).font = Font(bold=True)

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="producao_por_servidor_{date.today():%Y%m%d}.xlsx"'
    )
    wb.save(response)
    return response
